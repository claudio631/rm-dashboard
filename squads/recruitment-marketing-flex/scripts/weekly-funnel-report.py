#!/usr/bin/env python3
"""
Weekly OB Funnel Report Generator
Reads Tableau OB Funnel Custom Viewer .xlsx and optionally a requisitions CSV
to generate an HTML report with conditional formatting and open-req demand signals.

Usage:
    python3 weekly-funnel-report.py <xlsx_path> [--reqs <csv_path>] [--owner claudio|craig|all] [--output <path>]

Example:
    python3 weekly-funnel-report.py ~/Downloads/"OB Funnel Custom Viewer (2).xlsx" \
        --reqs ~/Downloads/requisitions-2026-03-19-123275.csv --owner claudio
"""

import openpyxl
import csv
import sys
import os
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# === CONFIGURATION ===

OWNER_MARKETS = {
    "claudio": [
        "Austin", "Houston", "Charlotte", "Fort Mill", "Orlando",
        "Las Vegas", "Reno", "Washington, D.C.", "Monroe", "Phoenix",
        "Logan Township"
    ],
    "craig": [
        "Columbus", "Cincinnati", "Hebron", "Chicago", "Atlanta",
        "Dallas", "Logan Township", "North Haven", "South Brunswick",
        "Hamilton", "Kearny", "Nashville", "Middleburg Heights",
        "Erlanger", "Newark", "Paulsboro", "McCarran"
    ]
}

STAGES = [
    'Worker Accounts Created',
    '1st Role Verified (# Workers)',
    '1st OB Task Completed (# Workers)',
    'Platform Verified (# Workers)',
    '"Ready to Book" Estimate (# Workers)',
    '1st Shift Booked (# Workers)',
    '1st Shift Completed (# Workers)'
]

SHORT_NAMES = ['Accts', 'Verif', 'OB Task', 'PlatV', 'Ready', 'Booked', 'Comp']

# 2025 benchmarks
BENCHMARKS = {
    'av': 56.0,   # Acct -> Verified
    'vo': 93.0,   # Verified -> OB Task
    'op': 69.1,   # OB Task -> PlatformV
    'pb': 50.4,   # PlatformV -> Booked
    'bc': 84.6,   # Booked -> Completed
    'cr': 15.4    # Overall
}

# Thresholds: (green_min, yellow_min) — below yellow_min = red
THRESHOLDS = {
    'av': (55, 40),
    'vo': (85, 75),
    'op': (65, 50),
    'pb': (40, 15),
    'bc': (80, 60),
    'cr': (15, 8)
}


def safe_div(a, b):
    return (a / b * 100) if b and b > 0 else None


def color_class(val, stage_key):
    if val is None:
        return "na"
    green_min, yellow_min = THRESHOLDS[stage_key]
    if val >= green_min:
        return "green"
    if val >= yellow_min:
        return "yellow"
    return "red"


def fmt_pct(val, stage_key):
    if val is None:
        return '<span class="na">—</span>'
    cls = color_class(val, stage_key)
    icon = {"green": "🟢", "yellow": "🟡", "red": "🔴", "na": ""}[cls]
    return f'<span class="{cls}">{icon} {val:.1f}%</span>'


def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find date columns and Grand Total
    date_cols = []
    grand_total_idx = None
    for i, h in enumerate(header):
        if h and "Grand Total" in str(h):
            grand_total_idx = i
        elif h and ("'" in str(h) or "20" in str(h)):
            date_cols.append((i, str(h)))

    dates = [str(h) for _, h in date_cols if h]
    week_start = dates[0] if dates else "?"
    week_end = dates[-1] if dates else "?"

    # Parse date headers into ISO format for JS (e.g. "Mar 01 '26" -> "2026-03-01")
    iso_dates = []
    for d in dates:
        try:
            dt = datetime.strptime(d.replace("'", "20"), "%b %d %Y")
            iso_dates.append(dt.strftime("%Y-%m-%d"))
        except ValueError:
            iso_dates.append(d)

    # Detect stage column: index 2 (old format) or 3 (new format with 3 Slice cols)
    stage_col = 2
    if header[2] and 'Slice' in str(header[2]):
        stage_col = 3

    data = {}
    daily_data = {}  # (client, location) -> stage -> [val_per_date]
    current_client = None
    current_location = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row = list(row)
        if row[0]:
            current_client = str(row[0]).strip()
        if row[1]:
            current_location = str(row[1]).strip()
        stage = row[stage_col]
        if not stage:
            continue
        stage = str(stage).strip()
        grand_total = row[grand_total_idx] if grand_total_idx and row[grand_total_idx] else 0

        key = (current_client, current_location)
        if key not in data:
            data[key] = {}
            daily_data[key] = {}
        data[key][stage] = int(grand_total) if grand_total else 0

        # Extract per-day values
        day_vals = []
        for col_idx, _ in date_cols:
            v = row[col_idx]
            day_vals.append(int(v) if v else 0)
        daily_data[key][stage] = day_vals

    return data, week_start, week_end, daily_data, iso_dates


# === REQUISITIONS PARSING ===

def parse_requisitions(csv_path):
    """Parse requisitions CSV and return a lookup dict keyed by (city, client_normalized)."""
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Only active demand
    active_statuses = ('open', 'auto-paused', 'draft')
    active = [r for r in rows if r.get('status', '').strip() in active_statuses]

    # Build lookup: city_key -> client_key -> {reqs, rsvps, target, titles, statuses}
    lookup = defaultdict(lambda: defaultdict(lambda: {
        'reqs': 0, 'rsvps': 0, 'target': 0, 'titles': set(), 'statuses': set()
    }))

    for r in active:
        loc = r.get('location', '').strip()
        client = r.get('client', '').strip()
        city = loc.split(',')[0].strip()
        city_key = _normalize_city(city)
        client_key = _normalize_client(client)

        entry = lookup[city_key][client_key]
        entry['reqs'] += 1
        try:
            entry['rsvps'] += int(r.get('rsvps', 0) or 0)
        except ValueError:
            pass
        try:
            entry['target'] += int(r.get('target_rsvps', 0) or 0)
        except ValueError:
            pass
        entry['titles'].add(r.get('job_title', '').strip())
        entry['statuses'].add(r.get('status', '').strip())

    return lookup


def _normalize_city(city):
    """Normalize city names for fuzzy matching between funnel and requisition data."""
    city = city.lower().strip()
    # Handle common variations
    mappings = {
        'washington, d.c.': 'washington',
        'washington, d.c': 'washington',
        'washington, dc': 'washington',
        'middleburg hts': 'middleburg heights',
        'fort mill': 'fort mill',
    }
    for k, v in mappings.items():
        if k in city:
            return v
    return city


def _normalize_client(client):
    """Normalize client names for fuzzy matching."""
    client = client.lower().strip()
    # Strip common suffixes
    for suffix in [', inc.', ', inc', ' inc.', ' inc', ', llc', ' llc',
                   ' corporation', ' corp', ' co.', ' co']:
        if client.endswith(suffix):
            client = client[:-len(suffix)]
    return client


def get_req_info(req_lookup, location, client):
    """Look up requisition info for a location/client combo.

    Also merges in 'Indeed Flex Application' requisitions for the same location,
    as these represent real demand signals for that market.
    """
    city_key = _normalize_city(location.split(',')[0].strip() if ',' in location else location)
    client_key = _normalize_client(client)
    ifx_key = _normalize_client('Indeed Flex Application')

    matched_info = None

    if city_key in req_lookup:
        city_data = req_lookup[city_key]
        # Exact match
        if client_key in city_data:
            matched_info = city_data[client_key]
        else:
            # Fuzzy: check if client_key is contained in any req client or vice versa
            for req_client_key, info in city_data.items():
                if client_key in req_client_key or req_client_key in client_key:
                    matched_info = info
                    break

        # Merge in Indeed Flex Application reqs for this location (skip if already matched to IFx)
        if ifx_key in city_data and client_key != ifx_key:
            ifx_info = city_data[ifx_key]
            if matched_info is None:
                matched_info = {'reqs': 0, 'rsvps': 0, 'target': 0,
                                'titles': set(), 'statuses': set()}
            matched_info = {
                'reqs': matched_info['reqs'] + ifx_info['reqs'],
                'rsvps': matched_info['rsvps'] + ifx_info['rsvps'],
                'target': matched_info['target'] + ifx_info['target'],
                'titles': matched_info['titles'] | ifx_info['titles'],
                'statuses': matched_info['statuses'] | ifx_info['statuses'],
            }

    return matched_info


def get_market_req_totals(req_lookup, location):
    """Get total requisitions for an entire market (all clients)."""
    city_key = _normalize_city(location.split(',')[0].strip() if ',' in location else location)
    if city_key not in req_lookup:
        return None

    total = {'reqs': 0, 'rsvps': 0, 'target': 0, 'titles': set(), 'statuses': set()}
    for client_key, info in req_lookup[city_key].items():
        total['reqs'] += info['reqs']
        total['rsvps'] += info['rsvps']
        total['target'] += info['target']
        total['titles'].update(info['titles'])
        total['statuses'].update(info['statuses'])

    return total if total['reqs'] > 0 else None


# === REPORT BUILDING ===

def build_report(data, markets, owner_name, week_start, week_end):
    market_data = {}
    grand_total = {s: 0 for s in STAGES}

    for loc in sorted(set(m for m in markets)):
        loc_total = {s: 0 for s in STAGES}
        loc_clients = {}

        for (client, location), stages in data.items():
            if location != loc:
                continue
            loc_clients[client] = {}
            for s in STAGES:
                val = stages.get(s, 0)
                loc_clients[client][s] = val
                loc_total[s] += val
                grand_total[s] += val

        if loc_clients:
            market_data[loc] = {'total': loc_total, 'clients': loc_clients}

    sorted_markets = sorted(market_data.items(),
                            key=lambda x: x[1]['total'][STAGES[0]], reverse=True)

    return sorted_markets, grand_total


def _fmt_req_cell(req_info, platv=0, is_total=False):
    """Format the requisition column cell with conditional coloring."""
    if req_info is None:
        if platv > 0:
            return '<span class="red">❌ No reqs</span>'
        return '<span class="na">—</span>'

    reqs = req_info['reqs']
    rsvps = req_info['rsvps']
    target = req_info['target']
    fill_pct = safe_div(rsvps, target) if target else None
    titles = sorted(req_info.get('titles', set()) - {''})
    title_str = ', '.join(t[:20] for t in titles[:3])

    if reqs == 0:
        if platv > 0:
            return '<span class="red">❌ No reqs</span>'
        return '<span class="na">—</span>'

    # Color based on fill rate
    if fill_pct is not None and fill_pct >= 80:
        cls = "yellow"
        icon = "⚠️"
        note = "near full"
    elif fill_pct is not None and fill_pct < 30:
        cls = "green"
        icon = "✅"
        note = "needs workers"
    else:
        cls = "green"
        icon = "✅"
        note = ""

    fill_str = f" ({fill_pct:.0f}%)" if fill_pct is not None else ""
    detail = f'<span class="{cls}">{icon} {reqs} reqs</span>'
    sub = f'<br><span style="font-size:10px;color:#666">{rsvps}/{target} RSVPs{fill_str}</span>'
    if title_str and not is_total:
        sub += f'<br><span style="font-size:9px;color:#999">{title_str}</span>'

    return detail + sub


def generate_html(sorted_markets, grand_total, owner_name, week_start, week_end, req_lookup=None,
                   daily_data=None, iso_dates=None, markets=None):
    gt = [grand_total[s] for s in STAGES]
    total_accts = gt[0]
    total_comp = gt[6]
    total_platv = gt[3]
    total_booked = gt[5]
    verified_not_booking = total_platv - total_booked

    has_reqs = req_lookup is not None
    colspan = "15" if has_reqs else "14"

    rows_html = ""

    # Grand total row
    a, v, ob, pv, rd, bk, co = gt
    req_cell = ""
    if has_reqs:
        # Sum all reqs across all markets in this report
        total_req_info = {'reqs': 0, 'rsvps': 0, 'target': 0, 'titles': set(), 'statuses': set()}
        for loc, _ in sorted_markets:
            mri = get_market_req_totals(req_lookup, loc)
            if mri:
                total_req_info['reqs'] += mri['reqs']
                total_req_info['rsvps'] += mri['rsvps']
                total_req_info['target'] += mri['target']
        req_cell = _fmt_req_cell(total_req_info, pv, is_total=True) if total_req_info['reqs'] > 0 else _fmt_req_cell(None, pv)

    rows_html += _make_row(
        f"<strong>{owner_name.upper()} TOTAL</strong>",
        a, v, ob, pv, rd, bk, co, req_cell=req_cell, is_total=True, has_reqs=has_reqs
    )
    rows_html += f'<tr class="spacer"><td colspan="{colspan}"></td></tr>'

    # Per market
    for loc, md in sorted_markets:
        t = md['total']
        a, v, ob, pv, rd, bk, co = [t[s] for s in STAGES]

        mkt_req_cell = ""
        if has_reqs:
            mri = get_market_req_totals(req_lookup, loc)
            mkt_req_cell = _fmt_req_cell(mri, pv, is_total=True)

        rows_html += _make_row(
            f"<strong>{loc.upper()}</strong>", a, v, ob, pv, rd, bk, co,
            req_cell=mkt_req_cell, is_market=True, has_reqs=has_reqs
        )

        sorted_clients = sorted(md['clients'].items(),
                                key=lambda x: x[1].get(STAGES[0], 0), reverse=True)
        for client, cs in sorted_clients:
            ca = cs.get(STAGES[0], 0)
            if ca == 0:
                continue
            cv, cob, cpv, crd, cbk, cco = [cs.get(s, 0) for s in STAGES[1:]]
            short_name = client if len(client) <= 35 else client[:33] + '..'

            client_req_cell = ""
            if has_reqs:
                cri = get_req_info(req_lookup, loc, client)
                client_req_cell = _fmt_req_cell(cri, cpv)

            rows_html += _make_row(
                f"&nbsp;&nbsp;- {short_name}", ca, cv, cob, cpv, crd, cbk, cco,
                req_cell=client_req_cell, has_reqs=has_reqs
            )

        rows_html += f'<tr class="spacer"><td colspan="{colspan}"></td></tr>'

    # Build HTML
    import json as _json
    today = date.today().strftime("%Y-%m-%d")
    cr_total = safe_div(total_comp, total_accts)
    cr_display = f"{cr_total:.1f}%" if cr_total else "0.0%"

    markets_with_comp = sum(1 for _, md in sorted_markets if md['total'][STAGES[6]] > 0)
    total_markets = len(sorted_markets)

    req_header_col = '<th>Open Reqs</th>' if has_reqs else ''
    req_data_source = " + Requisitions CSV" if has_reqs else ""

    # Build JSON data for JS date-range filtering
    js_daily = {}
    markets_set = set(markets) if markets else set()
    for (client, loc), stage_data in (daily_data or {}).items():
        if markets_set and loc not in markets_set:
            continue
        key = f"{client}|{loc}"
        js_daily[key] = {}
        for stage_name, vals in stage_data.items():
            if stage_name in STAGES:
                si = STAGES.index(stage_name)
                js_daily[key][si] = vals

    # Pre-compute req info for each market+client (req data is date-independent)
    js_req = {}
    if has_reqs:
        total_req = {'reqs': 0, 'rsvps': 0, 'target': 0}
        for loc, md in sorted_markets:
            mri = get_market_req_totals(req_lookup, loc)
            if mri:
                js_req[f"{loc}|__market__"] = {'r': mri['reqs'], 's': mri['rsvps'], 't': mri['target']}
                total_req['reqs'] += mri['reqs']
                total_req['rsvps'] += mri['rsvps']
                total_req['target'] += mri['target']
            for client in md['clients']:
                cri = get_req_info(req_lookup, loc, client)
                if cri:
                    js_req[f"{loc}|{client}"] = {'r': cri['reqs'], 's': cri['rsvps'], 't': cri['target']}
        js_req['__total__'] = {'r': total_req['reqs'], 's': total_req['rsvps'], 't': total_req['target']}

    daily_json = _json.dumps(js_daily, separators=(',', ':'))
    dates_json = _json.dumps(iso_dates or [], separators=(',', ':'))
    req_json = _json.dumps(js_req, separators=(',', ':'))

    # Build market structure for JS: [{loc, clients: [{name, shortName}]}] sorted by accts desc
    js_market_struct = []
    for loc, md in sorted_markets:
        clients_sorted = sorted(md['clients'].items(),
                                key=lambda x: x[1].get(STAGES[0], 0), reverse=True)
        cl = [{'n': c, 's': c if len(c) <= 35 else c[:33] + '..'} for c, cs in clients_sorted]
        js_market_struct.append({'loc': loc, 'clients': cl})
    markets_json = _json.dumps(js_market_struct, separators=(',', ':'))

    title_str = owner_name.title() + "'s Markets" if owner_name.lower() not in ('all markets', 'all') else 'All Markets'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Weekly OB Funnel Report — {title_str} ({week_start} to {week_end})</title>
<style>
    @media print {{
        body {{ margin: 0; font-size: 9pt; }}
        .page-break {{ page-break-before: always; }}
        @page {{ size: landscape; margin: 1cm; }}
        .date-filter {{ display: none !important; }}
    }}
    body {{
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif;
        max-width: 1500px;
        margin: 0 auto;
        padding: 20px;
        color: #1a1a1a;
        font-size: 13px;
        line-height: 1.4;
    }}
    h1 {{
        font-size: 22px;
        border-bottom: 3px solid #0066cc;
        padding-bottom: 8px;
        margin-bottom: 5px;
    }}
    h2 {{
        font-size: 16px;
        color: #0066cc;
        margin-top: 25px;
        border-bottom: 1px solid #ddd;
        padding-bottom: 4px;
    }}
    h3 {{ font-size: 14px; margin-top: 15px; }}
    .meta {{ color: #666; font-size: 12px; margin-bottom: 20px; }}
    .date-filter {{
        background: #eef3f8;
        border: 1px solid #c4d4e4;
        border-radius: 6px;
        padding: 10px 18px;
        margin: 12px 0;
        display: flex;
        align-items: center;
        gap: 12px;
        flex-wrap: wrap;
    }}
    .date-filter label {{
        font-weight: 600;
        font-size: 13px;
        color: #2c3e50;
    }}
    .date-filter input[type="date"] {{
        padding: 5px 10px;
        border: 1px solid #aab;
        border-radius: 4px;
        font-size: 13px;
        font-family: inherit;
    }}
    .date-filter button {{
        padding: 6px 16px;
        background: #0066cc;
        color: white;
        border: none;
        border-radius: 4px;
        font-size: 13px;
        cursor: pointer;
        font-weight: 600;
    }}
    .date-filter button:hover {{ background: #0052a3; }}
    .date-filter .preset {{
        padding: 4px 10px;
        background: #fff;
        border: 1px solid #aab;
        border-radius: 4px;
        font-size: 12px;
        cursor: pointer;
        color: #333;
    }}
    .date-filter .preset:hover {{ background: #dde5ee; }}
    .date-filter .preset.active {{ background: #0066cc; color: white; border-color: #0066cc; }}
    .date-filter .range-info {{
        font-size: 12px;
        color: #555;
        margin-left: auto;
    }}
    .summary-box {{
        background: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 6px;
        padding: 15px 20px;
        margin: 15px 0;
    }}
    .kpi-grid {{
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
        margin: 15px 0;
    }}
    .kpi-card {{
        background: #fff;
        border: 1px solid #dee2e6;
        border-radius: 6px;
        padding: 12px;
        text-align: center;
    }}
    .kpi-card .value {{ font-size: 24px; font-weight: bold; }}
    .kpi-card .label {{ font-size: 11px; color: #666; margin-top: 4px; }}
    .kpi-card.red {{ border-left: 4px solid #dc3545; }}
    .kpi-card.yellow {{ border-left: 4px solid #ffc107; }}
    .kpi-card.green {{ border-left: 4px solid #28a745; }}
    #funnel-table {{
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        margin: 10px 0;
        font-size: 12px;
    }}
    #funnel-table th {{
        background: #2c3e50;
        color: white;
        padding: 8px 6px;
        text-align: center;
        font-size: 11px;
        font-weight: 600;
        white-space: nowrap;
        position: sticky;
        top: 0;
        z-index: 20;
        cursor: help;
    }}
    #funnel-table th:first-child {{ text-align: left; min-width: 200px; position: sticky; left: 0; z-index: 30; background: #2c3e50; }}
    #funnel-table th[data-tip] {{ cursor: help; position: relative; }}
    #col-tooltip {{
        display: none;
        position: fixed;
        background: #1a1a1a;
        color: #fff;
        font-size: 12px;
        font-weight: 400;
        padding: 10px 14px;
        border-radius: 6px;
        z-index: 200;
        min-width: 240px;
        max-width: 340px;
        line-height: 1.6;
        box-shadow: 0 4px 16px rgba(0,0,0,0.35);
        pointer-events: none;
        white-space: pre-line;
    }}
    #funnel-table td {{
        padding: 5px 6px;
        border-bottom: 1px solid #eee;
        text-align: center;
        vertical-align: top;
    }}
    #funnel-table td:first-child {{ text-align: left; position: sticky; left: 0; z-index: 5; background: inherit; }}
    #funnel-table tr:not(.total-row):not(.market-row):not(.spacer):not(.spacer-after-total) td:first-child {{ background: #fff; }}
    #funnel-table tr.market-row td:first-child {{ background: #f0f4f8; }}
    #funnel-table tr.total-row td:first-child {{ background: #2c3e50; z-index: 15; }}
    #funnel-table tr.spacer td:first-child, #funnel-table tr.spacer-after-total td:first-child {{ background: #fff; }}
    #funnel-table td:last-child {{ text-align: left; min-width: 140px; }}
    tr.market-row {{ background: #f0f4f8; font-weight: 600; }}
    tr.market-row td {{ border-bottom: 1px solid #ccc; }}
    tr.total-row {{ background: #2c3e50; color: white; font-weight: 700; }}
    tr.total-row td {{ position: sticky; top: 30px; z-index: 10; background: #2c3e50; }}
    tr.total-row td {{ border-bottom: 2px solid #1a252f; }}
    tr.total-row .green, tr.total-row .yellow, tr.total-row .red, tr.total-row .num, tr.total-row .na {{ color: white; }}
    tr.total-row td {{ color: white; }}
    tr.total-row td span {{ color: white !important; }}
    tr.spacer-after-total td {{ border: none; height: 8px; background: white; position: sticky; top: 61px; z-index: 9; }}
    tr.spacer td {{ border: none; height: 8px; background: white; }}
    tr:hover:not(.spacer):not(.spacer-after-total):not(.total-row) {{ background: #f5f7fa; }}
    .green {{ color: #28a745; font-weight: 600; }}
    .yellow {{ color: #b8860b; font-weight: 600; }}
    .red {{ color: #dc3545; font-weight: 600; }}
    .na {{ color: #999; }}
    .num {{ color: #333; }}
    .legend-table {{ max-width: 700px; }}
    .legend-table td {{ padding: 4px 10px; }}
    .req-legend {{ max-width: 600px; margin-top: 10px; }}
    .req-legend td {{ padding: 3px 8px; font-size: 12px; }}
    .footer {{
        margin-top: 30px;
        padding-top: 10px;
        border-top: 1px solid #ddd;
        color: #999;
        font-size: 11px;
    }}
</style>
</head>
<body>

<h1>Weekly OB Funnel Report — {title_str}</h1>
<div class="meta">
    <strong>Period:</strong> <span id="period-display">{week_start} — {week_end}</span> &nbsp;|&nbsp;
    <strong>Generated:</strong> {today} &nbsp;|&nbsp;
    <strong>Source:</strong> OB Funnel Custom Viewer (Tableau){req_data_source} &nbsp;|&nbsp;
    <strong>Agent:</strong> Fiona (@funnel-specialist)
</div>

<div class="date-filter">
    <label>Date Range:</label>
    <input type="date" id="date-start">
    <span>to</span>
    <input type="date" id="date-end">
    <button onclick="applyDateRange()">Apply</button>
    <span style="color:#999">|</span>
    <button class="preset" onclick="setPreset('7d')">Last 7d</button>
    <button class="preset" onclick="setPreset('14d')">Last 14d</button>
    <button class="preset" onclick="setPreset('30d')" id="preset-30d">Full Range</button>
    <span class="range-info" id="range-info"></span>
</div>

<div id="kpi-grid" class="kpi-grid"></div>

<div id="summary-box" class="summary-box"></div>

<h2>Conditional Formatting</h2>
<table class="legend-table">
<tr><th>Stage</th><th>🟢 Good</th><th>🟡 Watch</th><th>🔴 Critical</th><th>2025 Benchmark</th></tr>
<tr><td>Acct &rarr; Verified</td><td class="green">&ge; 55%</td><td class="yellow">40-54%</td><td class="red">&lt; 40%</td><td>56.0%</td></tr>
<tr><td>Verified &rarr; OB Task</td><td class="green">&ge; 85%</td><td class="yellow">75-84%</td><td class="red">&lt; 75%</td><td>93.0%</td></tr>
<tr><td>OB Task &rarr; PlatformV</td><td class="green">&ge; 65%</td><td class="yellow">50-64%</td><td class="red">&lt; 50%</td><td>69.1%</td></tr>
<tr><td>PlatformV &rarr; Booked</td><td class="green">&ge; 40%</td><td class="yellow">15-39%</td><td class="red">&lt; 15%</td><td>50.4%</td></tr>
<tr><td>Booked &rarr; Completed</td><td class="green">&ge; 80%</td><td class="yellow">60-79%</td><td class="red">&lt; 60%</td><td>84.6%</td></tr>
<tr><td>Overall CR%</td><td class="green">&ge; 15%</td><td class="yellow">8-14.9%</td><td class="red">&lt; 8%</td><td>15.4%</td></tr>
</table>

{"<table class='req-legend'><tr><th colspan='2'>Open Reqs Column</th></tr><tr><td>✅ N reqs</td><td>Active requisitions — demand exists, workers can book</td></tr><tr><td>⚠️ N reqs</td><td>Reqs near full (RSVPs &ge; 80% of target)</td></tr><tr><td>❌ No reqs</td><td>No open requisitions — verified workers CANNOT book (red flag if PlatV &gt; 0)</td></tr></table>" if has_reqs else ""}

<h2>Full Funnel — By Location & Client</h2>
<table id="funnel-table">
<thead>
<tr>
    <th data-tip="Market location or client name">Market / Client</th>
    <th data-tip="Accounts Created&#10;Workers who created an account on the Indeed Flex platform">Accts</th>
    <th data-tip="1st Role Verified&#10;Workers who completed identity verification and got their first role verified">Verif</th>
    <th data-tip="Accounts to Verified %&#10;Formula: Verified / Accts x 100&#10;Benchmark: 56.0%&#10;Green >= 55% | Yellow >= 40% | Red < 40%">A&rarr;V</th>
    <th data-tip="1st OB Task Completed&#10;Workers who completed their first onboarding task (training, compliance)">OB Task</th>
    <th data-tip="Verified to OB Task %&#10;Formula: OB Task / Verified x 100&#10;Benchmark: 93.0%&#10;Green >= 85% | Yellow >= 75% | Red < 75%">V&rarr;OB</th>
    <th data-tip="Platform Verified&#10;Workers fully verified on the platform, eligible to be booked for shifts">PlatV</th>
    <th data-tip="OB Task to Platform Verified %&#10;Formula: PlatV / OB Task x 100&#10;Benchmark: 69.1%&#10;Green >= 65% | Yellow >= 50% | Red < 50%">OB&rarr;PV</th>
    <th data-tip="Ready to Book Estimate&#10;Workers estimated as ready to be booked for their first shift">Ready</th>
    <th data-tip="1st Shift Booked&#10;Workers who got booked for their very first shift">Booked</th>
    <th data-tip="Platform Verified to Booked %&#10;Formula: Booked / PlatV x 100&#10;Benchmark: 50.4%&#10;Green >= 40% | Yellow >= 15% | Red < 15%">PV&rarr;Bk</th>
    <th data-tip="1st Shift Completed&#10;Workers who successfully completed their first shift">Comp</th>
    <th data-tip="Booked to Completed %&#10;Formula: Completed / Booked x 100&#10;Benchmark: 84.6%&#10;Green >= 80% | Yellow >= 60% | Red < 60%">Bk&rarr;Co</th>
    <th data-tip="Overall Conversion Rate&#10;Formula: Completed / Accts Created x 100&#10;End-to-end funnel efficiency&#10;Benchmark: 15.4%&#10;Green >= 15% | Yellow >= 8% | Red < 8%">CR%</th>
    {'<th data-tip="Open Requisitions (FHS)&#10;Active job requisitions (open + auto-paused + draft)&#10;Shows RSVPs / Target and fill %">Open Reqs</th>' if has_reqs else ''}
</tr>
</thead>
<tbody id="funnel-tbody">
{rows_html}
</tbody>
</table>

<div class="footer">
    Report generated by Fiona (@funnel-specialist) &middot; Template: weekly-funnel-report-tmpl.md &middot; Squad: recruitment-marketing-flex<br>
    Data source: OB Funnel Custom Viewer (Tableau){req_data_source} &middot; Benchmarks: 2025 full-year baseline (52,987 accounts &rarr; 8,146 completed = 15.4%)
</div>

<script>
// === Embedded data ===
const DAILY = {daily_json};
const DATES = {dates_json};
const MARKETS = {markets_json};
const REQ = {req_json};
const HAS_REQS = {'true' if has_reqs else 'false'};
const OWNER = "{owner_name.upper()}";
const COLSPAN = {colspan};
const THRESHOLDS = {{av:[55,40],vo:[85,75],op:[65,50],pb:[40,15],bc:[80,60],cr:[15,8]}};

// === Helpers ===
function safeDiv(a,b) {{ return b>0 ? a/b*100 : null; }}

function fmtPct(val, key, forceWhite) {{
    if (val===null) return '<span class="na">&mdash;</span>';
    const [gMin,yMin] = THRESHOLDS[key];
    let cls = val>=gMin ? 'green' : val>=yMin ? 'yellow' : 'red';
    const icon = cls==='green' ? '🟢' : cls==='yellow' ? '🟡' : '🔴';
    return `<span class="${{cls}}">${{icon}} ${{val.toFixed(1)}}%</span>`;
}}

function fmtReq(reqInfo, platv, isTotal) {{
    if (!reqInfo || reqInfo.r===0) {{
        return platv>0 ? '<span class="red">❌ No reqs</span>' : '<span class="na">&mdash;</span>';
    }}
    const fillPct = reqInfo.t>0 ? reqInfo.s/reqInfo.t*100 : null;
    let cls='green', icon='✅';
    if (fillPct!==null && fillPct>=80) {{ cls='yellow'; icon='⚠️'; }}
    const fillStr = fillPct!==null ? ` (${{fillPct.toFixed(0)}}%)` : '';
    let h = `<span class="${{cls}}">${{icon}} ${{reqInfo.r}} reqs</span>`;
    h += `<br><span style="font-size:10px;color:#666">${{reqInfo.s}}/${{reqInfo.t}} RSVPs${{fillStr}}</span>`;
    return h;
}}

function sumRange(arr, si, ei) {{
    let s=0; for(let i=si;i<=ei;i++) s+=(arr[i]||0); return s;
}}

// === Compute and render ===
function computeAndRender(startIdx, endIdx) {{
    // Sum daily values for each client+location within the date range
    // stage indices: 0=Accts,1=Verif,2=OBTask,3=PlatV,4=Ready,5=Booked,6=Comp
    const clientData = {{}};  // "client|loc" -> [a,v,ob,pv,rd,bk,co]
    for (const [key, stages] of Object.entries(DAILY)) {{
        const vals = [0,0,0,0,0,0,0];
        for (let si=0; si<7; si++) {{
            if (stages[si]) vals[si] = sumRange(stages[si], startIdx, endIdx);
        }}
        clientData[key] = vals;
    }}

    // Build market aggregates, re-sort by accts descending
    const marketAgg = [];
    for (const mkt of MARKETS) {{
        const loc = mkt.loc;
        const mTotal = [0,0,0,0,0,0,0];
        const mClients = [];
        for (const cl of mkt.clients) {{
            const key = `${{cl.n}}|${{loc}}`;
            const vals = clientData[key] || [0,0,0,0,0,0,0];
            for (let i=0;i<7;i++) mTotal[i]+=vals[i];
            mClients.push({{name:cl.n, short:cl.s, vals}});
        }}
        // Re-sort clients by accts descending
        mClients.sort((a,b)=>b.vals[0]-a.vals[0]);
        marketAgg.push({{loc, total:mTotal, clients:mClients}});
    }}
    // Re-sort markets by accts descending
    marketAgg.sort((a,b)=>b.total[0]-a.total[0]);

    // Grand total
    const gt = [0,0,0,0,0,0,0];
    for (const m of marketAgg) for(let i=0;i<7;i++) gt[i]+=m.total[i];

    // Render KPIs
    const cr = safeDiv(gt[6],gt[0]);
    const crDisp = cr!==null ? cr.toFixed(1)+'%' : '0.0%';
    const vnb = gt[3]-gt[5];
    const crClass = (cr||0)<8?'red':(cr||0)<15?'yellow':'green';
    document.getElementById('kpi-grid').innerHTML = `
        <div class="kpi-card ${{crClass}}"><div class="value">${{crDisp}}</div><div class="label">Overall CR% (benchmark: 15.4%)</div></div>
        <div class="kpi-card"><div class="value">${{gt[0]}}</div><div class="label">Accounts Created</div></div>
        <div class="kpi-card"><div class="value">${{gt[6]}}</div><div class="label">Shifts Completed</div></div>
        <div class="kpi-card"><div class="value">${{vnb}}</div><div class="label">Verified but NOT Booking</div></div>`;

    // Render summary
    const vnbPct = gt[3]>0 ? (vnb/gt[3]*100).toFixed(0) : 0;
    const mktsComp = marketAgg.filter(m=>m.total[6]>0).length;
    document.getElementById('summary-box').innerHTML =
        `<strong>Portfolio snapshot:</strong> ${{gt[0]}} accounts &rarr; ${{gt[1]}} verified &rarr; ${{gt[3]}} platform verified &rarr; ${{gt[5]}} booked &rarr; ${{gt[6]}} completed`+
        `<br><strong>Markets with completions:</strong> ${{mktsComp}} of ${{marketAgg.length}}`+
        ` &nbsp;|&nbsp; <strong>Verified workers not booking:</strong> ${{vnb}} (${{vnbPct}}% of pipeline wasted)`;

    // Render table
    let html = '';
    const reqTd = (h) => HAS_REQS ? `<td>${{h}}</td>` : '';
    function makeRow(label, v, rowClass, reqHtml) {{
        const [a,vr,ob,pv,rd,bk,co] = v;
        const av=safeDiv(vr,a),vo=safeDiv(ob,vr),op=safeDiv(pv,ob),pb=safeDiv(bk,pv),bc=safeDiv(co,bk),crr=safeDiv(co,a);
        const cls = rowClass ? ` class="${{rowClass}}"` : '';
        return `<tr${{cls}}><td>${{label}}</td>`+
            `<td class="num">${{a}}</td>`+
            `<td class="num">${{vr}}</td><td>${{fmtPct(av,'av')}}</td>`+
            `<td class="num">${{ob}}</td><td>${{fmtPct(vo,'vo')}}</td>`+
            `<td class="num">${{pv}}</td><td>${{fmtPct(op,'op')}}</td>`+
            `<td class="num">${{rd}}</td>`+
            `<td class="num">${{bk}}</td><td>${{fmtPct(pb,'pb')}}</td>`+
            `<td class="num">${{co}}</td><td>${{fmtPct(bc,'bc')}}</td>`+
            `<td>${{fmtPct(crr,'cr')}}</td>`+
            reqTd(reqHtml||'') + `</tr>\\n`;
    }}

    // Grand total row (sticky)
    const totalReq = HAS_REQS ? fmtReq(REQ['__total__'], gt[3], true) : '';
    html += makeRow(`<strong>${{OWNER}} TOTAL</strong>`, gt, 'total-row', totalReq);
    html += `<tr class="spacer-after-total"><td colspan="${{COLSPAN}}"></td></tr>`;

    // Per market
    for (const m of marketAgg) {{
        const mReq = HAS_REQS ? fmtReq(REQ[`${{m.loc}}|__market__`], m.total[3], true) : '';
        html += makeRow(`<strong>${{m.loc.toUpperCase()}}</strong>`, m.total, 'market-row', mReq);
        for (const cl of m.clients) {{
            if (cl.vals[0]===0) continue;
            const cReq = HAS_REQS ? fmtReq(REQ[`${{m.loc}}|${{cl.name}}`], cl.vals[3]) : '';
            html += makeRow(`&nbsp;&nbsp;- ${{cl.short}}`, cl.vals, '', cReq);
        }}
        html += `<tr class="spacer"><td colspan="${{COLSPAN}}"></td></tr>`;
    }}

    document.getElementById('funnel-tbody').innerHTML = html;
}}

// === Date range controls ===
const dateStart = document.getElementById('date-start');
const dateEnd = document.getElementById('date-end');

if (DATES.length > 0) {{
    dateStart.min = DATES[0]; dateStart.max = DATES[DATES.length-1];
    dateEnd.min = DATES[0]; dateEnd.max = DATES[DATES.length-1];
    dateStart.value = DATES[0];
    dateEnd.value = DATES[DATES.length-1];
}}

function getDateIndices() {{
    const s = dateStart.value, e = dateEnd.value;
    let si=0, ei=DATES.length-1;
    for (let i=0;i<DATES.length;i++) {{ if(DATES[i]>=s){{ si=i; break; }} }}
    for (let i=DATES.length-1;i>=0;i--) {{ if(DATES[i]<=e){{ ei=i; break; }} }}
    return [si, ei];
}}

function applyDateRange() {{
    const [si, ei] = getDateIndices();
    const days = ei-si+1;
    document.getElementById('range-info').textContent = `${{days}} day${{days!==1?'s':''}} selected`;
    document.getElementById('period-display').textContent = `${{dateStart.value}} — ${{dateEnd.value}}`;
    // Clear preset highlights
    document.querySelectorAll('.preset').forEach(b=>b.classList.remove('active'));
    computeAndRender(si, ei);
}}

function setPreset(key) {{
    const last = DATES.length-1;
    let si=0;
    if (key==='7d') si=Math.max(0, last-6);
    else if (key==='14d') si=Math.max(0, last-13);
    else si=0;
    dateStart.value = DATES[si];
    dateEnd.value = DATES[last];
    document.querySelectorAll('.preset').forEach(b=>b.classList.remove('active'));
    if (key==='7d') document.querySelectorAll('.preset')[0].classList.add('active');
    else if (key==='14d') document.querySelectorAll('.preset')[1].classList.add('active');
    else document.getElementById('preset-30d').classList.add('active');
    applyDateRange();
}}

// Initial render with full range
document.getElementById('preset-30d').classList.add('active');
applyDateRange();

// === Column tooltips ===
(function() {{
    const tip = document.getElementById('col-tooltip');
    document.querySelectorAll('#funnel-table th[data-tip]').forEach(th => {{
        th.addEventListener('mouseenter', e => {{
            const text = th.getAttribute('data-tip');
            tip.textContent = text;
            tip.style.display = 'block';
            const rect = th.getBoundingClientRect();
            let left = rect.left + rect.width/2 - 140;
            if (left < 8) left = 8;
            if (left + 340 > window.innerWidth) left = window.innerWidth - 350;
            tip.style.left = left + 'px';
            tip.style.top = (rect.bottom + 6) + 'px';
        }});
        th.addEventListener('mouseleave', () => {{
            tip.style.display = 'none';
        }});
    }});
}})();
</script>

<div id="col-tooltip"></div>

</body>
</html>"""

    return html


def _make_row(label, a, v, ob, pv, rd, bk, co, req_cell="", is_market=False, is_total=False, has_reqs=False):
    av = safe_div(v, a)
    vo = safe_div(ob, v)
    op = safe_div(pv, ob)
    pb = safe_div(bk, pv)
    bc = safe_div(co, bk)
    cr = safe_div(co, a)

    row_class = ""
    if is_total:
        row_class = ' class="total-row"'
    elif is_market:
        row_class = ' class="market-row"'

    req_td = f"<td>{req_cell}</td>" if has_reqs else ""

    return f"""<tr{row_class}>
    <td>{label}</td>
    <td class="num">{a}</td>
    <td class="num">{v}</td><td>{fmt_pct(av, 'av')}</td>
    <td class="num">{ob}</td><td>{fmt_pct(vo, 'vo')}</td>
    <td class="num">{pv}</td><td>{fmt_pct(op, 'op')}</td>
    <td class="num">{rd}</td>
    <td class="num">{bk}</td><td>{fmt_pct(pb, 'pb')}</td>
    <td class="num">{co}</td><td>{fmt_pct(bc, 'bc')}</td>
    <td>{fmt_pct(cr, 'cr')}</td>
    {req_td}
</tr>\n"""


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Weekly OB Funnel Report Generator')
    parser.add_argument('xlsx_path', help='Path to OB Funnel Custom Viewer .xlsx')
    parser.add_argument('--reqs', default=None,
                        help='Path to requisitions CSV (adds Open Reqs column)')
    parser.add_argument('--owner', default='claudio', choices=['claudio', 'craig', 'all'],
                        help='Market owner to filter (default: claudio)')
    parser.add_argument('--client', default=None,
                        help='Filter by client name (case-insensitive substring match, e.g. --client CORT)')
    parser.add_argument('--output', default=None,
                        help='Output HTML path (default: docs/reports/weekly-funnel-YYYY-MM-DD.html)')
    args = parser.parse_args()

    xlsx_path = os.path.expanduser(args.xlsx_path)
    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading funnel: {xlsx_path}")
    data, week_start, week_end, daily_data, iso_dates = parse_xlsx(xlsx_path)

    req_lookup = None
    if args.reqs:
        reqs_path = os.path.expanduser(args.reqs)
        if not os.path.exists(reqs_path):
            print(f"Warning: Requisitions file not found: {reqs_path}, skipping")
        else:
            print(f"Reading requisitions: {reqs_path}")
            req_lookup = parse_requisitions(reqs_path)
            total_active = sum(sum(info['reqs'] for info in city.values()) for city in req_lookup.values())
            print(f"Active requisitions loaded: {total_active}")

    # Filter by client if specified
    if args.client:
        client_filter = args.client.strip().lower()
        filtered_data = {k: v for k, v in data.items() if client_filter in k[0].lower()}
        filtered_daily = {k: v for k, v in daily_data.items() if client_filter in k[0].lower()}
        print(f"Client filter: '{args.client}' — {len(filtered_data)} entries (from {len(data)})")
        data = filtered_data
        daily_data = filtered_daily

    if args.owner == 'all':
        markets = list(set(loc for (_, loc) in data.keys()))
        owner_name = "All Markets"
    else:
        markets = OWNER_MARKETS[args.owner]
        owner_name = args.owner.title()

    print(f"Owner: {owner_name} ({len(markets)} markets)")
    sorted_markets, grand_total = build_report(data, markets, owner_name, week_start, week_end)

    html = generate_html(sorted_markets, grand_total, owner_name, week_start, week_end, req_lookup,
                         daily_data=daily_data, iso_dates=iso_dates, markets=markets)

    if args.output:
        output_path = args.output
    else:
        project_root = Path(__file__).resolve().parent.parent.parent.parent
        reports_dir = project_root / "docs" / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        today = date.today().strftime("%Y-%m-%d")
        client_suffix = f"-{args.client.lower()}" if args.client else ""
        output_path = str(reports_dir / f"weekly-funnel-{args.owner}{client_suffix}-{today}.html")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"Report saved: {output_path}")
    print(f"Open in browser and use Cmd+P / Ctrl+P to print as PDF")

    return output_path


if __name__ == '__main__':
    main()
