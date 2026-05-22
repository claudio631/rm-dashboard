#!/usr/bin/env python3
"""
Ad Test Tracker — Multi-channel spreadsheet generator
Scout (@ad-test-specialist) · recruitment-marketing-flex squad
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.chart import BarChart, Reference
import os

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_BLUE   = "0F2B5B"
GREEN       = "00A878"
RED         = "E63946"
AMBER       = "F4A261"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F7FA"
MID_GRAY    = "D0D5DD"
DARK_GRAY   = "343A40"
LIGHT_BLUE  = "E8EFF8"
LIGHT_GREEN = "D4F4EC"
LIGHT_RED   = "FADDE0"
LIGHT_AMBER = "FDE8D4"
VERY_LIGHT  = "FAFBFD"

# ── Style Helpers ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK_GRAY, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(sides="all", color=MID_GRAY, style="thin"):
    s = Side(style=style, color=color)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    if sides == "outer":
        return Border(left=s, right=s, top=s, bottom=s)
    return Border()

def style_cell(ws, row, col, value="", bg=None, bold=False, color=DARK_GRAY,
               size=10, h="left", v="center", wrap=False, italic=False,
               num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, color=color, size=size, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    c.border = border()
    if num_format:
        c.number_format = num_format
    return c

def header_row(ws, row, labels, col_start=1, bg=DARK_BLUE, color=WHITE,
               size=10, bold=True, height=None):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=color, size=size)
        c.alignment = align(h="center", v="center", wrap=True)
        c.border = border()
    if height:
        ws.row_dimensions[row].height = height

def section_header(ws, row, label, col_start, col_end, bg=DARK_BLUE,
                   color=WHITE, size=11, height=22):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.fill = fill(bg)
    c.font = font(bold=True, color=color, size=size)
    c.alignment = align(h="left", v="center")
    c.border = border(sides="outer")
    ws.row_dimensions[row].height = height

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — TEST REGISTRY (master list of all tests)
# ══════════════════════════════════════════════════════════════════════════════

def build_registry(wb):
    ws = wb.create_sheet("📋 Test Registry", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "AD TEST REGISTRY — Indeed Flex RM Team"
    c.fill = fill(DARK_BLUE)
    c.font = font(bold=True, color=WHITE, size=16)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:P2")
    c = ws["A2"]
    c.value = "Master log of all ad tests across channels · @ad-test-specialist (Scout) · Updated manually after each *log-metrics run"
    c.fill = fill(LIGHT_BLUE)
    c.font = font(italic=True, color=DARK_BLUE, size=9)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[2].height = 16

    # Headers
    headers = [
        "Test ID", "Test Name", "Channel", "Test Type", "Variable",
        "Owner", "Start Date", "End Date", "Status",
        "Control (A)", "Variant (B)",
        "Primary KPI", "A Result", "B Result", "Lift %", "Winner"
    ]
    header_row(ws, 3, headers, height=30)

    # Sample rows
    sample_rows = [
        ["google-nashville-headline-20260601", "Nashville Headline Test",
         "Google", "Ad Copy", "Headline 1 phrasing",
         "Claudio", "2026-06-01", "2026-06-14", "🟡 Running",
         "Earn up to $XX/hr in Nashville", "Nashville Warehouse Jobs — Apply Now",
         "CPA", "", "", "", ""],
        ["indeed-dallas-title-20260601", "Dallas Job Title Test",
         "Indeed", "Job Title", "Hourly rate visible vs hidden",
         "Claudio", "2026-06-01", "2026-06-10", "🔵 Draft",
         "Warehouse Worker — $18-22/hr", "Warehouse Worker — Immediate Openings",
         "Apply Start Rate", "", "", "", ""],
        ["reddit-nashville-format-20260525", "Nashville Format Test",
         "Reddit", "Ad Format", "Image vs Free-form text",
         "Claudio", "2026-05-25", "2026-06-05", "🟡 Running",
         "Standard Image Ad", "Free-form Text Ad",
         "CPR", "", "", "", ""],
    ]

    status_colors = {
        "🟡 Running": LIGHT_AMBER,
        "🔵 Draft": LIGHT_BLUE,
        "✅ Complete": LIGHT_GREEN,
        "⏸ Paused": LIGHT_GRAY,
        "❌ Cancelled": LIGHT_RED,
    }

    for ri, row_data in enumerate(sample_rows):
        r = 4 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row_data):
            cell_bg = bg
            if ci == 8:  # Status column
                cell_bg = status_colors.get(val, bg)
            elif ci == 14 and val:  # Lift %
                try:
                    lift = float(val.replace("%", ""))
                    cell_bg = LIGHT_GREEN if lift > 0 else LIGHT_RED
                except:
                    pass
            style_cell(ws, r, ci + 1, val, bg=cell_bg,
                       h="center" if ci in [2, 3, 6, 7, 8, 14, 15] else "left",
                       wrap=ci in [1, 9, 10])

    # Empty rows for data entry
    for ri in range(len(sample_rows), 50):
        r = 4 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        for ci in range(16):
            style_cell(ws, r, ci + 1, "", bg=bg)

    # Data validation — Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"🔵 Draft,🟡 Running,⏸ Paused,✅ Complete,❌ Cancelled"',
        showDropDown=False
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = "I4:I53"

    # Data validation — Channel dropdown
    dv_channel = DataValidation(
        type="list",
        formula1='"Google,Indeed,Reddit,Meta,TikTok"',
        showDropDown=False
    )
    ws.add_data_validation(dv_channel)
    dv_channel.sqref = "C4:C53"

    # Column widths
    set_col_widths(ws, {
        1: 38, 2: 28, 3: 10, 4: 16, 5: 22, 6: 12,
        7: 12, 8: 12, 9: 14, 10: 30, 11: 30,
        12: 18, 13: 12, 14: 12, 15: 10, 16: 14,
    })

    freeze(ws, "C4")
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDER — Generic per-channel test detail sheet
# ══════════════════════════════════════════════════════════════════════════════

CHANNEL_CONFIGS = {
    "Google": {
        "icon": "🔵",
        "color": DARK_BLUE,
        "metrics": [
            ("Impressions", "Number"),
            ("Clicks", "Number"),
            ("CTR", "Percentage"),
            ("Spend ($)", "Currency"),
            ("CPC ($)", "Currency"),
            ("Conversions", "Number"),
            ("Conv Rate", "Percentage"),
            ("CPA ($)", "Currency"),
        ],
        "test_types": "Ad Copy · Creative · Audience · Bid Strategy · Landing Page · Match Type · RSA Pins",
        "min_note": "Min: 30 conversions per variant · Min runtime: 7 days · Use 'Do not optimize' rotation",
        "kpis": "Primary KPI: CPA  |  Secondary: Conv Rate, CTR",
        "best_practices": [
            "One variable changed between control and variant",
            "Both ads in same ad group (equal auction exposure)",
            "Ad rotation: 'Do not optimize' during test period",
            "RSA tests: pin Headline 1 vs unpinned to isolate variable",
            "Exclude branded terms from test campaigns",
            "Check impression share before declaring low-volume winner",
        ]
    },
    "Indeed": {
        "icon": "🔷",
        "color": "003087",
        "metrics": [
            ("Impressions", "Number"),
            ("Clicks", "Number"),
            ("CTR", "Percentage"),
            ("Spend ($)", "Currency"),
            ("CPC ($)", "Currency"),
            ("Apply Starts", "Number"),
            ("Apply Start Rate", "Percentage"),
            ("Cost/Apply Start ($)", "Currency"),
        ],
        "test_types": "Job Title · Job Description · Salary Display · Sponsored Level · CTA Phrasing",
        "min_note": "Min: 50 apply starts per variant · Min runtime: 7 days (avoid day-of-week bias)",
        "kpis": "Primary KPI: Apply Start Rate  |  Secondary: Cost/Apply Start, CTR",
        "best_practices": [
            "Same job category and location across all variants",
            "Same sponsored budget level — only change the test variable",
            "Test salary displayed vs hidden (significant CTR impact)",
            "Test job title phrasing: role clarity vs urgency ('Now Hiring' vs specific title)",
            "Run minimum 7 days — Monday traffic ≠ Friday traffic",
            "Apply Starts (not platform applies) is the valid conversion metric",
        ]
    },
    "Reddit": {
        "icon": "🟠",
        "color": "FF4500",
        "metrics": [
            ("Impressions", "Number"),
            ("Clicks", "Number"),
            ("CTR", "Percentage"),
            ("Spend ($)", "Currency"),
            ("CPC ($)", "Currency"),
            ("Results", "Number"),
            ("CPR ($)", "Currency"),
            ("eCPM ($)", "Currency"),
        ],
        "test_types": "Ad Format · Creative · Headline · Subreddit Targeting · Audience · Free-form vs Image",
        "min_note": "Min: 50 results per variant · Min runtime: 5 days · Watch frequency closely",
        "kpis": "Primary KPI: CPR  |  Secondary: CTR, Results volume",
        "best_practices": [
            "Free-form text ads consistently outperform image-only on Reddit",
            "Match community tone — no corporate jargon, no stock images",
            "Same subreddit targeting across variants (isolate one variable)",
            "Frequency cap: pause if >4 impressions per user (creative fatigue)",
            "Test local city subreddits vs job-specific subreddits separately",
            "Upvote count is visible — strong copy earns social proof organically",
        ]
    },
    "Meta": {
        "icon": "🔵",
        "color": "1877F2",
        "metrics": [
            ("Impressions", "Number"),
            ("Reach", "Number"),
            ("Frequency", "Number"),
            ("Clicks", "Number"),
            ("CTR", "Percentage"),
            ("Spend ($)", "Currency"),
            ("CPC ($)", "Currency"),
            ("Results", "Number"),
            ("CPR ($)", "Currency"),
            ("ROAS", "Number"),
        ],
        "test_types": "Creative · Headline · Primary Text · Audience · Placement · Format · CTA Button",
        "min_note": "Min: 50 results per variant · Min runtime: 7 days · Check audience overlap before launch",
        "kpis": "Primary KPI: CPR  |  Secondary: ROAS, CTR, Frequency",
        "best_practices": [
            "Check audience overlap — overlapping audiences dilute test validity",
            "Use Campaign Budget Optimization only after establishing baselines",
            "Video: test hook (first 3 seconds) as the primary variable",
            "Static vs video: run as separate tests, not within same ad set",
            "Frequency cap awareness tests at 3×/week max to avoid fatigue",
            "Broad audience + strong creative outperforms narrow targeting in 2026",
        ]
    },
    "TikTok": {
        "icon": "⚫",
        "color": "010101",
        "metrics": [
            ("Impressions", "Number"),
            ("Reach", "Number"),
            ("Video Views", "Number"),
            ("VTR (25%)", "Percentage"),
            ("VTR (100%)", "Percentage"),
            ("Clicks", "Number"),
            ("CTR", "Percentage"),
            ("Spend ($)", "Currency"),
            ("CPC ($)", "Currency"),
            ("CPA ($)", "Currency"),
        ],
        "test_types": "Hook (First 3s) · Caption · CTA · Audience · Bid Strategy · Spark Ad vs Regular",
        "min_note": "Min: 500 video views per variant · Min runtime: 5 days · Hook is the #1 test variable",
        "kpis": "Primary KPI: CPA  |  Secondary: VTR (100%), CTR",
        "best_practices": [
            "Hook test = change ONLY first 3 seconds, keep rest identical",
            "Always review sound-on AND sound-off versions before launch",
            "Spark Ads (boosting organic posts) often outperform regular ads",
            "Caption CTA test separately from creative — don't mix variables",
            "TikTok algorithm needs ~500 views to exit learning phase per variant",
            "Native-looking content outperforms polished production on this platform",
        ]
    },
}


def build_channel_sheet(wb, channel_name):
    cfg = CHANNEL_CONFIGS[channel_name]
    icon = cfg["icon"]
    accent = cfg["color"]
    metrics = cfg["metrics"]

    ws = wb.create_sheet(f"{icon} {channel_name}")
    ws.sheet_view.showGridLines = False

    # ── Title ──
    last_col = 20
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1,
                value=f"{channel_name.upper()} ADS — Test Performance Tracker")
    c.fill = fill(accent)
    c.font = font(bold=True, color=WHITE, size=15)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=1,
                value=f"Test Types: {cfg['test_types']}   ·   {cfg['kpis']}")
    c.fill = fill(LIGHT_BLUE)
    c.font = font(italic=True, color=DARK_BLUE, size=9)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[2].height = 14

    # ── SECTION A: Test Setup ──────────────────────────────────────────────
    section_header(ws, 3, "A  TEST SETUP", 1, last_col, bg=accent)

    setup_fields = [
        ("Test ID",          "e.g. google-nashville-headline-20260601"),
        ("Test Name",        "Short descriptive name"),
        ("Test Type",        cfg["test_types"].split("·")[0].strip()),
        ("Variable",         "The ONE thing being changed between A and B"),
        ("Owner",            "Your name"),
        ("Start Date",       "YYYY-MM-DD"),
        ("End Date",         "YYYY-MM-DD (minimum runtime enforced)"),
        ("Status",           "🔵 Draft / 🟡 Running / ✅ Complete"),
        ("Hypothesis",       "If X, then Y will change by Z% because W"),
        ("Success Metric",   cfg["kpis"].split("|")[0].replace("Primary KPI:", "").strip()),
        ("Min. Threshold",   cfg["min_note"]),
    ]

    label_col, value_col = 1, 3
    for ri, (label, placeholder) in enumerate(setup_fields):
        r = 4 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        # Label
        c = ws.cell(row=r, column=label_col, value=label)
        c.fill = fill(LIGHT_BLUE)
        c.font = font(bold=True, color=DARK_BLUE, size=10)
        c.alignment = align(h="right", v="center")
        c.border = border()
        ws.merge_cells(start_row=r, start_column=label_col,
                       end_row=r, end_column=2)
        # Value cell
        ws.merge_cells(start_row=r, start_column=value_col,
                       end_row=r, end_column=10)
        c2 = ws.cell(row=r, column=value_col, value=placeholder if ri > 7 else "")
        c2.fill = fill(bg)
        c2.font = font(italic=(ri > 7), color=MID_GRAY if ri > 7 else DARK_GRAY,
                       size=10)
        c2.alignment = align(h="left", v="center")
        c2.border = border()
        ws.row_dimensions[r].height = 18

    # ── SECTION B: Variants ────────────────────────────────────────────────
    B_ROW = 4 + len(setup_fields) + 1
    section_header(ws, B_ROW, "B  VARIANTS — Creative & Copy Links", 1,
                   last_col, bg=accent)

    variant_headers = ["", "Variant Label", "Description", "Creative Link",
                       "Copy Doc Link", "Landing Page", "UTM Tag", "Ad / Post ID"]
    vw = [2, 14, 25, 25, 25, 20, 22, 16]
    header_row(ws, B_ROW + 1, variant_headers, bg=DARK_GRAY)

    for vi, (vlabel, vbg) in enumerate([("A — Control", LIGHT_GREEN),
                                         ("B — Variant", LIGHT_AMBER)]):
        r = B_ROW + 2 + vi
        style_cell(ws, r, 1, "►" if vi == 1 else "◆", bg=vbg,
                   bold=True, h="center", color=DARK_BLUE)
        style_cell(ws, r, 2, vlabel, bg=vbg, bold=True, color=DARK_BLUE)
        for ci in range(2, 8):
            style_cell(ws, r, ci + 1, "", bg=WHITE if vi == 0 else VERY_LIGHT)
        ws.row_dimensions[r].height = 20

    # ── SECTION C: Metric Log ──────────────────────────────────────────────
    C_ROW = B_ROW + 5
    section_header(ws, C_ROW, "C  METRIC SNAPSHOTS — Log every 2–3 days while running",
                   1, last_col, bg=accent)

    metric_names = [m[0] for m in metrics]
    col_headers = (["Date", "Variant"] + metric_names + ["Notes"])
    header_row(ws, C_ROW + 1, col_headers, bg=DARK_GRAY)

    # 30 data rows (A + B pairs)
    for pair in range(15):
        for vi, (vlabel, vbg) in enumerate([("A — Control", LIGHT_GREEN),
                                             ("B — Variant", LIGHT_AMBER)]):
            r = C_ROW + 2 + pair * 2 + vi
            date_bg = LIGHT_BLUE if pair % 2 == 0 else WHITE
            style_cell(ws, r, 1, "", bg=date_bg)
            style_cell(ws, r, 2, vlabel, bg=vbg, bold=False, color=DARK_BLUE,
                       size=9)
            for ci in range(len(metric_names)):
                mt = metrics[ci][1]
                nf = None
                if mt == "Percentage":
                    nf = "0.00%"
                elif mt == "Currency":
                    nf = '$#,##0.00'
                style_cell(ws, r, ci + 3, "", bg=WHITE, num_format=nf)
            style_cell(ws, r, 3 + len(metric_names), "", bg=VERY_LIGHT)
            ws.row_dimensions[r].height = 16

    # ── SECTION D: Results Summary ─────────────────────────────────────────
    D_ROW = C_ROW + 2 + 30 + 1
    section_header(ws, D_ROW, "D  RESULTS SUMMARY — Fill when minimum thresholds reached",
                   1, last_col, bg=accent)

    result_headers = ["Metric", "Control (A)", "Variant (B)", "Lift %", "Winner ✓"]
    header_row(ws, D_ROW + 1, result_headers[:5], bg=DARK_GRAY)
    for ri, (metric, _) in enumerate(metrics):
        r = D_ROW + 2 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        style_cell(ws, r, 1, metric, bg=LIGHT_BLUE, bold=True, color=DARK_BLUE)
        for ci in range(4):
            style_cell(ws, r, ci + 2, "", bg=bg)
        ws.row_dimensions[r].height = 16

    # Significance row
    sig_r = D_ROW + 2 + len(metrics)
    ws.merge_cells(start_row=sig_r, start_column=1,
                   end_row=sig_r, end_column=2)
    style_cell(ws, sig_r, 1, "Statistical Significance", bg=LIGHT_AMBER,
               bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=sig_r, start_column=3,
                   end_row=sig_r, end_column=5)
    style_cell(ws, sig_r, 3, "Flag if <80% — need more data",
               bg=LIGHT_AMBER, italic=True, color=DARK_GRAY)

    # ── SECTION E: Winner & Insight ────────────────────────────────────────
    E_ROW = sig_r + 2
    section_header(ws, E_ROW, "E  WINNER DECLARATION & KEY INSIGHT", 1,
                   last_col, bg=accent)

    winner_fields = [
        "Winner Variant (A / B / Inconclusive)",
        "Primary Lift Achieved",
        "Declared On (date)",
        "Declared By",
        "Recommended Action (scale winner / kill loser / run follow-up test)",
        "Key Insight (one sentence, actionable for the whole team)",
        "Next Test Recommendation",
    ]
    for ri, label in enumerate(winner_fields):
        r = E_ROW + 1 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        style_cell(ws, r, 1, label, bg=LIGHT_BLUE, bold=True, color=DARK_BLUE)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=last_col)
        style_cell(ws, r, 5, "", bg=bg)
        ws.row_dimensions[r].height = 20

    # ── SECTION F: Best Practices ──────────────────────────────────────────
    F_ROW = E_ROW + 1 + len(winner_fields) + 1
    section_header(ws, F_ROW, f"F  {channel_name.upper()} TESTING BEST PRACTICES",
                   1, last_col, bg=DARK_GRAY)

    for ri, practice in enumerate(cfg["best_practices"]):
        r = F_ROW + 1 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        # Checkbox column
        c = ws.cell(row=r, column=1, value="☐")
        c.fill = fill(bg)
        c.font = font(size=12, color=GREEN)
        c.alignment = align(h="center", v="center")
        c.border = border()
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=last_col)
        style_cell(ws, r, 2, practice, bg=bg, color=DARK_GRAY, size=10, wrap=True)
        ws.row_dimensions[r].height = 18

    # ── Column widths ──────────────────────────────────────────────────────
    set_col_widths(ws, {
        1: 14,   # Date / label
        2: 16,   # Variant / value
        3: 16,   # Metric 1 / value
        4: 14,
        5: 14,
        6: 14,
        7: 14,
        8: 14,
        9: 14,
        10: 14,
        11: 14,
        12: 14,
        13: 14,
        14: 14,
        15: 14,
        16: 14,
        17: 14,
        18: 14,
        19: 14,
        20: 18,
    })

    freeze(ws, "C4")
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET — Benchmarks & Channel KPI Reference
# ══════════════════════════════════════════════════════════════════════════════

def build_benchmarks(wb):
    ws = wb.create_sheet("📐 Benchmarks & KPIs")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "CHANNEL BENCHMARKS & KPI REFERENCE — Indeed Flex Recruitment Marketing"
    c.fill = fill(DARK_BLUE)
    c.font = font(bold=True, color=WHITE, size=14)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "Use these as baseline expectations when reading test results. Lift = (Variant − Control) / Control × 100"
    c.fill = fill(LIGHT_BLUE)
    c.font = font(italic=True, color=DARK_BLUE, size=9)
    c.alignment = align(h="center")
    ws.row_dimensions[2].height = 14

    benchmarks = [
        # Channel, Metric, Good, Target, Weak, Unit, Notes
        ("Google",  "CTR",             "5%+",    "3–5%",   "<2%",    "%",  "Search campaigns — recruitment vertical"),
        ("Google",  "CPC",             "<$0.80", "$0.80–1.50","$1.50+","$", "Varies by market competitiveness"),
        ("Google",  "CPA",             "<$1.50", "$1.50–2.50","$2.50+","$", "App conversions + form fills"),
        ("Google",  "Conv Rate",       "50%+",   "30–50%", "<30%",   "%",  "Click to conversion"),
        ("Indeed",  "CTR",             "4%+",    "2–4%",   "<2%",    "%",  "Sponsored job listings"),
        ("Indeed",  "CPC",             "<$0.70", "$0.70–1.20","$1.20+","$","Sponsored clicks"),
        ("Indeed",  "Apply Start Rate","30%+",   "15–30%", "<15%",   "%",  "Clicks to apply start"),
        ("Indeed",  "Cost/Apply Start","<$3.00", "$3–6",   "$6+",    "$",  "Primary conversion metric"),
        ("Reddit",  "CTR",             "0.5%+",  "0.3–0.5%","<0.3%", "%",  "Paid placements"),
        ("Reddit",  "CPC",             "<$1.50", "$1.50–3.00","$3+",  "$",  "Varies by subreddit/audience"),
        ("Reddit",  "CPR",             "<$1.00", "$1–3",   "$3+",    "$",  "Cost per result (app event/lead)"),
        ("Meta",    "CTR",             "2%+",    "1–2%",   "<1%",    "%",  "Feed placements"),
        ("Meta",    "CPC",             "<$1.00", "$1–2",   "$2+",    "$",  "All placements blended"),
        ("Meta",    "CPR",             "<$3.00", "$3–8",   "$8+",    "$",  "Lead/apply events"),
        ("Meta",    "Frequency",       "1–3",    "3–5",    "5+",     "x",  "High freq = creative fatigue"),
        ("TikTok",  "CTR",             "1.5%+",  "0.8–1.5%","<0.8%", "%",  "In-feed ads"),
        ("TikTok",  "VTR (25%)",       "60%+",   "40–60%", "<40%",   "%",  "25% video completion"),
        ("TikTok",  "VTR (100%)",      "15%+",   "8–15%",  "<8%",    "%",  "Full video completion"),
        ("TikTok",  "CPC",             "<$1.00", "$1–2.50","$2.50+", "$",  "In-feed CPC"),
        ("TikTok",  "CPA",             "<$5.00", "$5–12",  "$12+",   "$",  "App install / lead"),
    ]

    headers = ["Channel", "Metric", "Good ✅", "Target 🟡", "Weak 🔴", "Unit",
               "Notes"]
    header_row(ws, 3, headers, bg=DARK_BLUE, height=24)

    last_channel = None
    for ri, row in enumerate(benchmarks):
        r = 4 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        channel_changed = row[0] != last_channel
        if channel_changed:
            last_channel = row[0]

        for ci, val in enumerate(row[:7]):
            cell_bg = bg
            if ci == 0:  # Channel
                ch_colors = {"Google": "E8EFF8", "Indeed": "E8F4FF",
                             "Reddit": "FFF0EA", "Meta": "E8F0FE", "TikTok": "F0F0F0"}
                cell_bg = ch_colors.get(val, bg)
            elif ci == 2:  # Good
                cell_bg = LIGHT_GREEN
            elif ci == 3:  # Target
                cell_bg = LIGHT_AMBER
            elif ci == 4:  # Weak
                cell_bg = LIGHT_RED

            style_cell(ws, r, ci + 1, val, bg=cell_bg,
                       bold=(ci == 0 and channel_changed),
                       h="center" if ci in [2, 3, 4, 5] else "left",
                       color=DARK_GRAY)
        ws.row_dimensions[r].height = 16

    set_col_widths(ws, {1: 10, 2: 20, 3: 12, 4: 14, 5: 12, 6: 8, 7: 40})
    freeze(ws, "C4")
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET — Test Ideas Backlog
# ══════════════════════════════════════════════════════════════════════════════

def build_backlog(wb):
    ws = wb.create_sheet("💡 Test Backlog")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "TEST IDEAS BACKLOG — Add ideas here, promote to a channel sheet when ready to run"
    c.fill = fill(DARK_BLUE)
    c.font = font(bold=True, color=WHITE, size=13)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[1].height = 28

    headers = ["Priority", "Channel", "Test Idea", "Hypothesis (brief)",
               "Expected KPI", "Estimated Lift", "Status", "Assigned To", "Notes"]
    header_row(ws, 2, headers, bg=DARK_GRAY, height=22)

    seed_ideas = [
        ("🔴 High", "Google", "App vs Search CPA comparison",
         "App campaigns deliver 2.7× better CPA — test budget reallocation",
         "CPA", "30–50% improvement", "🔵 Backlog", "Claudio", "Based on May review"),
        ("🔴 High", "Indeed", "Salary visible vs hidden (Dallas)",
         "Showing '$18-22/hr' in title increases apply start rate by 20%+",
         "Apply Start Rate", "15–25% lift", "🔵 Backlog", "Claudio", "Priority market"),
        ("🟡 Medium", "Reddit", "Free-form vs image creative",
         "Free-form text feels more authentic — expect higher CPR",
         "CPR", "40% improvement", "🔵 Backlog", "Claudio", "Nashville market"),
        ("🟡 Medium", "Google", "Reno keyword restructure",
         "Broad+phrase mix (Phoenix model) vs current — CPA $4.77 → <$2.50",
         "CPA", "47% reduction", "🔵 Backlog", "Claudio", "Reno currently 2.7× target"),
        ("🟢 Low", "Meta", "Static vs video creative",
         "Video hook drives higher CVR for warehouse roles",
         "CPR", "20% improvement", "🔵 Backlog", "", ""),
        ("🟢 Low", "TikTok", "Hook variant test — urgency vs curiosity",
         "Urgency hook ('We're hiring now') vs curiosity ('Here's what no one tells you about...')",
         "VTR (100%)", "25% lift", "🔵 Backlog", "", ""),
    ]

    priority_colors = {"🔴 High": LIGHT_RED, "🟡 Medium": LIGHT_AMBER, "🟢 Low": LIGHT_GREEN}

    for ri, row in enumerate(seed_ideas):
        r = 3 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row):
            cell_bg = priority_colors.get(val, bg) if ci == 0 else bg
            style_cell(ws, r, ci + 1, val, bg=cell_bg,
                       h="center" if ci in [0, 1, 6, 7] else "left",
                       wrap=ci in [2, 3, 8])
        ws.row_dimensions[r].height = 20

    # Empty rows
    for ri in range(len(seed_ideas), 50):
        r = 3 + ri
        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
        for ci in range(9):
            style_cell(ws, r, ci + 1, "", bg=bg)

    # Dropdowns
    dv_p = DataValidation(type="list",
                          formula1='"🔴 High,🟡 Medium,🟢 Low"')
    ws.add_data_validation(dv_p)
    dv_p.sqref = "A3:A52"

    dv_c = DataValidation(type="list",
                          formula1='"Google,Indeed,Reddit,Meta,TikTok,Cross-channel"')
    ws.add_data_validation(dv_c)
    dv_c.sqref = "B3:B52"

    dv_s = DataValidation(type="list",
                          formula1='"🔵 Backlog,🟡 Running,✅ Complete,❌ Cancelled"')
    ws.add_data_validation(dv_s)
    dv_s.sqref = "G3:G52"

    set_col_widths(ws, {1: 12, 2: 10, 3: 30, 4: 40, 5: 18, 6: 16,
                        7: 14, 8: 14, 9: 28})
    freeze(ws, "C3")
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

wb = Workbook()
# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

build_registry(wb)
for channel in ["Google", "Indeed", "Reddit", "Meta", "TikTok"]:
    build_channel_sheet(wb, channel)
build_benchmarks(wb)
build_backlog(wb)

# Tab colors
tab_colors = {
    "📋 Test Registry":  "0F2B5B",
    "🔵 Google":         "4285F4",
    "🔷 Indeed":         "003087",
    "🟠 Reddit":         "FF4500",
    "🔵 Meta":           "1877F2",
    "⚫ TikTok":         "010101",
    "📐 Benchmarks & KPIs": "00A878",
    "💡 Test Backlog":   "F4A261",
}
for sheet in wb.sheetnames:
    for key, color in tab_colors.items():
        if key in sheet or sheet in key:
            wb[sheet].sheet_properties.tabColor = color
            break

out_path = "/Users/claudio.santos/RM-Team-Ai/data/ad-test-tracker.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"✅ Saved: {out_path}")
print(f"   Sheets: {wb.sheetnames}")
