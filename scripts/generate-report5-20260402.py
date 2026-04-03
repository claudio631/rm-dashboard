#!/usr/bin/env python3
"""
Report 5: Revenue Requests Cross-Reference Dashboard — 2026-04-02
Generates an HTML dashboard cross-referencing Revenue Requests with FHS, Indeed (JobsCampaigns), and OB Funnel data.

CRITICAL: Every qualifying row in Revenue Requests = one row in output. NO merging/dedup.
First column = Status badge (O for Live, D for Declined)
Sort by submission date ascending (chronological).
"""

import csv
import re
import os
import sys
from collections import defaultdict
from datetime import datetime, date

import openpyxl

# =============================================================================
# FILE PATHS
# =============================================================================
REVENUE_CSV = os.path.expanduser("~/Downloads/US_Recruitment_Requests__us_ (11).csv")
FHS_CSV = os.path.expanduser("~/Downloads/requisitions-2026-04-02-495947.csv")
INDEED_CSV = os.path.expanduser("~/Downloads/JobsCampaigns_20260101_20260326.csv")
OB_FUNNEL_XLSX = os.path.expanduser("~/Downloads/OB Funnel Custom Viewer (24).xlsx")
OUTPUT_HTML = os.path.expanduser("~/RM-Team-Ai/docs/reports/revenue-requests-crossref-2026-04-02.html")
REPORT_DATE = "2026-04-02"

# =============================================================================
# CLIENT NAME NORMALIZATION MAPS
# =============================================================================
CLIENT_NAME_MAP = {
    "compass": ["culinaire", "culinaire international"],
    "dc flex": ["cort", "ontrac final mile"],
    "food glorious": ["culinaire"],
    "bon appetit": ["bon appetit management company, inc", "bon appetit"],
    "legends": ["legends hospitality"],
    "levy": ["levy restaurants"],
    "btx": ["btx global logistics"],
    "lettuce": ["lettuce entertain you enterprises, inc"],
    "soho house": ["soho house austin"],
    "austin club": ["austin catering"],
    "vestals": ["vestals catering"],
    "rhino": ["rhino staging"],
    "power stop": ["power stop"],
    "uplift": ["uplift desk"],
    "ryerson": ["ryerson"],
    "solaren": ["solaren risk management"],
    "hei hotels": ["merritt hospitality llc"],
    "hyatt place": ["merritt hospitality llc", "hyatt", "hei"],
    "hyatt": ["merritt hospitality llc"],
    "merritt": ["merritt hospitality llc"],
    "foxconn": ["foxconn"],
    "ctdi": ["ctdi"],
    "ingram": ["ingram content group"],
    "stord": ["stord, inc"],
    "ontrac": ["ontrac final mile"],
    "tennant": ["tennant solutions"],
    "continental battery": ["continental battery systems, inc."],
    "bath concepts": ["bath concepts"],
    "cort": ["cort"],
    "assurant": ["assurant, inc (hyla)", "assurant"],
    "sxsw": ["sxsw"],
    "culinaire": ["culinaire", "culinaire international"],
    "material handler": ["ctdi"],
    "need forklift": ["johnstone supply"],
    "eurest": ["eurest", "eurest usa"],
    "freshella": ["freshella catering", "freshella"],
    "flik": ["flik"],
}

# Metro/Location aliases
LOCATION_ALIASES = {
    "dfw": "dallas",
    "vegas": "las vegas",
    "bedford park": "chicago",
    "hodgkins": "chicago",
    "fort worth": "dallas",
    "lancaster": "dallas",
    "flower mound": "dallas",
    "haslet": "dallas",
    "lavergne": "nashville",
    "la vergne": "nashville",
    "mt. juliet": "nashville",
    "mount juliet": "nashville",
    "libertyville": "chicago",
    "grove city": "columbus",
    "mccarran": "las vegas",
    "middleburg heights": "columbus",
    "lockbourne": "columbus",
    "logan township": "new jersey",
    "paulsboro": "new jersey",
    "hebron": "erlanger",
    "west chester": "cincinnati",
    "washington": "washington, d.c.",
    "washington dc": "washington, d.c.",
    "washington, dc": "washington, d.c.",
    "washington d.c.": "washington, d.c.",
    "washington, d.c.": "washington, d.c.",
    "washington, d.c": "washington, d.c.",
    "bedford": "chicago",
    "arlington tx": "dallas",
    "arlington": "dallas",
    "kissimmee": "orlando",
    "desoto": "dallas",
    "plano": "dallas",
    "garland": "dallas",
    "mesquite": "dallas",
    "fairfield": "cincinnati",
    "cartersville": "atlanta",
    "erlanger": "cincinnati",
    "fort mill": "charlotte",
    "nw atlanta": "atlanta",
    "wimberley": "austin",
    "pasadena": "houston",
    "inland empire": "inland empire",
    "reno": "reno",
    "joliet": "chicago",
    "sparks": "reno",
    "hapeville": "atlanta",
    "carrolton": "dallas",
    "irving": "dallas",
}


def normalize_location(loc_str):
    """Normalize a location string to a canonical metro name."""
    if not loc_str:
        return ""
    loc = loc_str.lower().strip()

    # Check aliases FIRST on the raw string
    for alias, metro in LOCATION_ALIASES.items():
        if alias in loc:
            return metro

    # Check common metros on raw string
    for metro in ["chicago", "dallas", "austin", "nashville", "las vegas",
                  "orlando", "houston", "atlanta", "columbus", "cincinnati",
                  "charlotte", "phoenix", "reno", "newark", "new jersey",
                  "philadelphia", "cleveland", "san antonio", "inland empire",
                  "washington, d.c."]:
        if metro in loc:
            return metro

    # Remove state abbreviations and zip codes
    loc = re.sub(r',\s*[A-Za-z]{2}\s*\d{0,5}\s*$', '', loc_str.lower().strip())
    loc = loc.strip().rstrip(',').strip()

    for alias, metro in LOCATION_ALIASES.items():
        if alias in loc:
            return metro

    for metro in ["chicago", "dallas", "austin", "nashville", "las vegas",
                  "orlando", "houston", "atlanta", "columbus", "cincinnati",
                  "charlotte", "phoenix", "reno", "newark", "new jersey",
                  "philadelphia", "cleveland", "san antonio", "inland empire",
                  "washington, d.c."]:
        if metro in loc:
            return metro

    return loc


def normalize_client_for_match(client_short):
    """Get list of possible FHS/Indeed/Funnel client names."""
    c = client_short.lower().strip()
    for key, vals in CLIENT_NAME_MAP.items():
        if key == c or c.startswith(key):
            return vals
    return [c]


def extract_client_and_location(client_job_str):
    """Extract short client name and location from the Client-Job column."""
    s = client_job_str.strip()
    sl = s.lower()

    # Special cases
    if sl.startswith("need forklift"):
        return "Johnstone Supply", "Dallas"
    if sl == "material handler":
        return "CTDI", "Dallas"
    if sl.startswith("san antonio & austin"):
        return "CORT", "San Antonio/Austin"
    if sl.startswith("assurant"):
        return "Assurant", "Nashville"
    if sl == "foxconn":
        return "Foxconn", "Dallas"
    if sl == "rhino staging":
        return "Rhino", "Austin"
    if sl.startswith("sxsw"):
        return "SXSW", "Austin"
    if "vestals catering" in sl:
        return "Vestals", "Austin"
    if "obama foundation" in sl:
        return "Bon Appetit", "Chicago"
    if "dos equis pavilion" in sl and "legends" in sl:
        if "at&t" not in sl:
            return "Legends", "Dallas"
    if "at&t stadium" in sl and "legends" in sl:
        return "Legends", "Dallas"
    if "soldier field" in sl and "levy" in sl:
        return "Levy", "Chicago"
    if "barista" == sl:
        return "Barista", "Dallas"
    if sl.startswith("plano, texas") or sl == "plano, texas - dallas":
        return "Unknown", "Dallas"

    parts = re.split(r'\s*-\s*', s, maxsplit=3)
    client = parts[0].strip()

    # Remove leading parens like "(Culinaire)"
    if client.startswith("("):
        m = re.match(r'\(([^)]+)\)\s*(.*)', client)
        if m:
            client = m.group(2).strip() if m.group(2).strip() else m.group(1).strip()

    location = ""
    if len(parts) >= 2:
        location = parts[1].strip()

    loc = _extract_city(location, s)

    # Client cleanups
    cl = client.lower()
    client_map = {
        "cort": "CORT", "culinaire": "Culinaire", "food glorious": "Culinaire",
        "compass": "Compass", "bon appetit": "Bon Appetit", "dc flex": "DC Flex",
        "foxconn": "Foxconn", "power stop": "Power Stop", "legends": "Legends",
        "levy": "Levy", "ingram": "Ingram", "ctdi": "CTDI",
        "bath concepts": "Bath Concepts", "btx": "BTX", "solaren": "Solaren",
        "hei hotels": "HEI Hotels", "hyatt": "Hyatt Place", "rhino": "Rhino",
        "uplift": "Uplift", "ryerson": "Ryerson", "lettuce": "Lettuce",
        "soho": "Soho House", "austin club": "Austin Club", "vestals": "Vestals",
        "sxsw": "SXSW", "eurest": "Eurest", "freshella": "Freshella",
        "flik": "FLIK", "ncr voyix": "NCR Voyix", "stord": "Stord",
    }
    for key, name in client_map.items():
        if key in cl:
            client = name
            break

    return client, loc


def _extract_city(loc_part, full_str):
    """Try to extract a city name from a location part."""
    cities = {
        "chicago": "Chicago", "dallas": "Dallas", "dfw": "Dallas",
        "austin": "Austin", "nashville": "Nashville", "las vegas": "Las Vegas",
        "vegas": "Las Vegas", "orlando": "Orlando", "houston": "Houston",
        "atlanta": "Atlanta", "columbus": "Columbus", "cincinnati": "Cincinnati",
        "charlotte": "Charlotte", "phoenix": "Phoenix", "reno": "Reno",
        "newark": "Newark", "new jersey": "New Jersey",
        "san antonio": "San Antonio", "washington": "Washington, D.C.",
        "fort worth": "Dallas", "bedford park": "Chicago", "hodgkins": "Chicago",
        "la vergne": "Nashville", "lavergne": "Nashville", "grove city": "Columbus",
        "libertyville": "Chicago", "bedford": "Chicago", "haslet": "Dallas",
        "mt. juliet": "Nashville", "mount juliet": "Nashville",
        "carrolton": "Dallas", "irving": "Dallas",
    }
    check_str = full_str.lower()
    for key, city in cities.items():
        if key in check_str:
            return city

    if loc_part:
        cleaned = re.sub(r'\d+\s+\w+\s+(dr|drive|st|street|ave|blvd|rd|way)\b.*', '', loc_part, flags=re.IGNORECASE).strip()
        if cleaned:
            return cleaned
    return loc_part


# Market owners
MARKET_OWNERS = {
    "claudio": [
        "austin", "houston", "charlotte", "fort mill", "orlando",
        "las vegas", "reno", "washington, d.c.", "washington", "monroe", "phoenix",
        "logan township"
    ],
    "craig": [
        "columbus", "cincinnati", "hebron", "chicago", "atlanta",
        "dallas", "north haven", "south brunswick",
        "hamilton", "kearny", "nashville", "middleburg heights",
        "erlanger", "newark", "paulsboro", "mccarran"
    ],
}


def get_owner_by_market(location):
    loc_norm = normalize_location(location)
    for owner, markets in MARKET_OWNERS.items():
        for market in markets:
            if market in loc_norm or loc_norm in market:
                return owner.title()
    return ""


def parse_owner(email):
    owners = {
        "claudio.santos": "Claudio", "craig.freeman": "Craig",
        "megan.mccue": "Megan", "angela": "Angela",
        "lacey.henderson": "Lacey", "david.starkman": "David",
    }
    email_lower = email.lower().strip()
    for key, name in owners.items():
        if key in email_lower:
            return name
    m = re.match(r'([a-z]+)', email_lower)
    return m.group(1).title() if m else email


def parse_headcount(hc_str):
    s = hc_str.strip()
    if not s:
        return 0
    if s.upper() == "LFDTW":
        return "LFDTW"
    m = re.search(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 0


def has_shifts(shift_info):
    if not shift_info:
        return False
    s = shift_info.lower().strip()
    no_indicators = ["not posted", "tbd", "no posted", "shifts not posted",
                     "need to immediately start recruiting", "not yet",
                     "will be based on client posts"]
    for ind in no_indicators:
        if ind in s:
            return False
    if re.search(r'\d+\s*(am|pm|a\.m|p\.m)', s, re.IGNORECASE):
        return True
    if re.search(r'(monday|tuesday|wednesday|thursday|friday|saturday|sunday|mon|tue|wed|thu|fri|sat|sun|m-f)', s, re.IGNORECASE):
        return True
    if len(s) > 10:
        return True
    return False


def parse_currency(val_str):
    if not val_str or val_str.strip() == '-':
        return 0.0
    s = val_str.strip().replace('$', '').replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# =============================================================================
# LOAD DATA
# =============================================================================

def load_revenue_requests():
    """Load all Live and Declined rows from revenue requests CSV."""
    rows = []
    with open(REVENUE_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 12:
                continue
            status = row[2].strip()
            # Include Live and Declined only
            if "Live" not in status and "Declined" not in status:
                continue
            is_declined = "Declined" in status

            client_job = row[0].strip()
            owner_email = row[1].strip()
            role_col = row[5].strip()
            other_role = row[6].strip() if len(row) > 6 else ""
            shift_info = row[9].strip() if len(row) > 9 else ""
            start_date = row[10].strip() if len(row) > 10 else ""
            hc_str = row[11].strip() if len(row) > 11 else "0"
            submission_str = row[17].strip() if len(row) > 17 else ""

            # Parse submission date (format: "25/03/2026, 12:55" DD/MM/YYYY)
            submission_date = None
            if submission_str:
                try:
                    submission_date = datetime.strptime(submission_str.split(',')[0].strip(), '%d/%m/%Y').date()
                except ValueError:
                    pass

            client, location = extract_client_and_location(client_job)
            owner = get_owner_by_market(location) or parse_owner(owner_email)
            hc = parse_headcount(hc_str)

            role = role_col
            if role.lower() == "other" and other_role:
                role = other_role
            if not role:
                parts = re.split(r'\s*-\s*', client_job)
                if len(parts) >= 3:
                    role = parts[-1].strip()
                    if re.match(r'^\d', role) or len(role) > 60:
                        role = parts[2].strip() if len(parts) > 2 else ""

            shifts = has_shifts(shift_info)

            # Format start date as mmm/dd/yyyy
            formatted_date = start_date
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    dt = datetime.strptime(start_date, fmt)
                    formatted_date = dt.strftime('%b/%d/%Y')
                    break
                except ValueError:
                    continue

            rows.append({
                'client_job': client_job,
                'client': client,
                'location': location,
                'owner': owner,
                'role': role,
                'hc': hc,
                'start_date': formatted_date,
                'shifts': shifts,
                'shift_info': shift_info,
                'is_declined': is_declined,
                'submission_date': submission_date,
            })

    return rows


def load_fhs_requisitions():
    """Load FHS requisitions."""
    rows = []
    with open(FHS_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 13:
                continue
            last_updated_str = row[0].strip()
            job_title = row[3].strip().lower() if len(row) > 3 else ""
            client = row[5].strip()
            location = row[6].strip()
            status = row[12].strip()
            rsvps = int(row[10]) if row[10].strip().isdigit() else 0

            try:
                req_date = datetime.strptime(last_updated_str[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                req_date = None

            loc_norm = normalize_location(location)
            rows.append({
                'client': client.lower().strip(),
                'loc_norm': loc_norm,
                'job_title': job_title,
                'status': status,
                'rsvps': rsvps,
                'date': req_date,
            })
    return rows


def load_indeed_campaigns():
    """Load Indeed campaigns from JobsCampaigns CSV.
    This file has ALL campaigns (no status filter). Match by campaign name keywords.
    Columns: Campaign, Spend, Applies, Job Count, etc.
    """
    rows = []
    with open(INDEED_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            campaign_name = row.get('Campaign', '').strip()
            spend = parse_currency(row.get('Spend', '0'))
            applies = 0
            applies_str = row.get('Applies', '0').strip()
            try:
                applies = int(float(applies_str))
            except (ValueError, TypeError):
                applies = 0

            # Skip "Jobs not in a campaign" and empty
            if not campaign_name or campaign_name.lower() == "jobs not in a campaign":
                continue

            rows.append({
                'name': campaign_name,
                'name_lower': campaign_name.lower(),
                'spend': spend,
                'applies': applies,
            })
    return rows


def load_ob_funnel():
    """Load OB Funnel data from Excel.
    Col 0 = Client, Col 1 = Location, Col 2 = "Total" type, Col 3 = Metric
    Date cols from 4, Grand Total = last col (index 34).
    Only use rows where Col 2 = "Total".
    Key metrics: Worker Accounts Created, 1st Role Verified, "Ready to Book" Estimate.
    """
    data = defaultdict(lambda: {'created': 0, 'verified': 0, 'rtb': 0})

    wb = openpyxl.load_workbook(OB_FUNNEL_XLSX, data_only=True)
    ws = wb.active

    current_client = None
    grand_total_col = ws.max_column  # Last column is Grand Total

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and str(row[0]).strip():
            current_client = str(row[0]).strip()
        if not current_client:
            continue

        col2 = row[2]
        if col2 is None or str(col2).strip() != "Total":
            continue

        metric = row[3]
        if not metric:
            continue
        metric_str = str(metric).strip()

        location = str(row[1]).strip() if row[1] else ""
        grand_total = row[grand_total_col - 1]  # 0-indexed
        if grand_total is None:
            grand_total = 0
        try:
            val = int(float(grand_total))
        except (ValueError, TypeError):
            val = 0

        loc_norm = normalize_location(location)
        key = (current_client.lower().strip(), loc_norm)

        if "Worker Accounts Created" in metric_str:
            data[key]['created'] += val
        elif "1st Role Verified" in metric_str:
            data[key]['verified'] += val
        elif '"Ready to Book" Estimate' in metric_str or "Ready to Book" in metric_str:
            data[key]['rtb'] += val

    wb.close()
    return data


# =============================================================================
# CROSS-REFERENCE MATCHING
# =============================================================================

def match_fhs(client_short, location, fhs_rows, submission_date=None, role=None):
    """Find FHS data matching a revenue request."""
    possible_clients = normalize_client_for_match(client_short)
    loc_norm = normalize_location(location)

    total = {'count': 0, 'rsvps': 0}

    for fhs_row in fhs_rows:
        client_match = False
        for pc in possible_clients:
            pc_lower = pc.lower().strip()
            if (pc_lower in fhs_row['client'] or fhs_row['client'] in pc_lower or
                    _fuzzy_client_match(pc_lower, fhs_row['client'])):
                client_match = True
                break
        if not client_match:
            continue
        if not _location_match(loc_norm, fhs_row['loc_norm']):
            continue

        if role:
            if not fhs_row['job_title']:
                continue
            if not _role_match(role, fhs_row['job_title']):
                continue

        if fhs_row['status'] in ('open', 'auto-paused'):
            total['count'] += 1

        if submission_date and fhs_row['date']:
            if fhs_row['date'] >= submission_date:
                total['rsvps'] += fhs_row['rsvps']
        elif not submission_date:
            total['rsvps'] += fhs_row['rsvps']

    return total


def match_indeed(client_short, location, indeed_rows, role=None):
    """Find Indeed campaigns matching a revenue request by keywords in campaign name."""
    loc_norm = normalize_location(location)

    # Build keywords from client, location, role
    client_keywords = _get_client_keywords(client_short)
    location_keywords = _get_location_keywords(location, loc_norm)

    total = {'count': 0, 'spend': 0.0}

    for ind in indeed_rows:
        name = ind['name_lower']

        # Check client keywords match
        client_match = any(kw in name for kw in client_keywords)
        if not client_match:
            continue

        # Check location keywords match
        loc_match = any(kw in name for kw in location_keywords) if location_keywords else True
        if not loc_match:
            continue

        total['count'] += 1
        total['spend'] += ind['spend']

    return total


def _get_client_keywords(client_short):
    """Build keyword list for client matching in campaign names."""
    c = client_short.lower().strip()
    kw = [c]

    keyword_map = {
        "cort": ["cort"],
        "culinaire": ["culinaire", "food glorious"],
        "compass": ["culinaire", "compass"],
        "dc flex": ["dc flex", "indeed flex"],
        "bon appetit": ["bon appetit"],
        "legends": ["legends"],
        "levy": ["levy"],
        "power stop": ["power stop", "powerstop"],
        "foxconn": ["foxconn"],
        "ctdi": ["ctdi"],
        "ingram": ["ingram"],
        "bath concepts": ["bath concepts"],
        "rhino": ["rhino"],
        "solaren": ["solaren"],
        "hyatt place": ["hyatt"],
        "hei hotels": ["hei", "hyatt"],
        "eurest": ["eurest"],
        "freshella": ["freshella"],
        "ryerson": ["ryerson"],
        "johnstone supply": ["johnstone"],
        "stord": ["stord"],
        "merritt": ["merritt", "hei"],
    }
    for key, vals in keyword_map.items():
        if key in c or c in key:
            return vals
    return kw


def _get_location_keywords(location, loc_norm):
    """Build keyword list for location matching in campaign names."""
    if not location and not loc_norm:
        return []
    keywords = []
    if loc_norm:
        keywords.append(loc_norm)
    loc_lower = location.lower().strip() if location else ""

    # Add state abbreviations
    loc_to_state = {
        "chicago": "il", "dallas": "tx", "austin": "tx", "nashville": "tn",
        "las vegas": "nv", "orlando": "fl", "houston": "tx", "atlanta": "ga",
        "columbus": "oh", "cincinnati": "oh", "charlotte": "nc", "phoenix": "az",
    }
    if loc_norm in loc_to_state:
        keywords.append(loc_to_state[loc_norm])

    # Add common variations
    if loc_norm == "dallas":
        keywords.extend(["dfw", "fort worth", "arlington"])
    elif loc_norm == "chicago":
        keywords.extend(["bedford park", "hodgkins", "libertyville"])
    elif loc_norm == "nashville":
        keywords.extend(["mt. juliet", "la vergne", "lavergne"])
    elif loc_norm == "columbus":
        keywords.extend(["grove city", "lockbourne"])

    return keywords


OB_FUNNEL_CLIENT_OVERRIDES = {
    "dc flex": ["cort", "ontrac final mile"],
}


def match_ob_funnel(client_short, location, ob_data, role=None):
    """Find OB Funnel data matching a revenue request."""
    possible_clients = normalize_client_for_match(client_short)
    override_clients = OB_FUNNEL_CLIENT_OVERRIDES.get(client_short.lower().strip(), [])
    if override_clients:
        possible_clients = possible_clients + override_clients
    loc_norm = normalize_location(location)

    total = {'created': 0, 'verified': 0, 'rtb': 0}

    for (ob_client, ob_loc), vals in ob_data.items():
        # Location match
        if not _location_match(loc_norm, ob_loc):
            continue

        # Client match
        client_match = False
        for pc in possible_clients:
            pc_lower = pc.lower().strip()
            if (pc_lower in ob_client or ob_client in pc_lower or
                    _fuzzy_client_match(pc_lower, ob_client)):
                client_match = True
                break
        if not client_match:
            continue

        total['created'] += vals['created']
        total['verified'] += vals['verified']
        total['rtb'] += vals['rtb']

    return total


def _role_match(rev_role, fhs_job_title):
    a = rev_role.lower().strip()
    b = fhs_job_title.lower().strip()
    if not a or not b:
        return True
    if a == b:
        return True
    if a in b or b in a:
        return True
    role_aliases = {
        'food preps': ['prep cook', 'prep'],
        'prep cook': ['food preps', 'prep'],
        'loader/crew': ['loader', 'crew', 'lc', 'loader / crew'],
        'loader': ['loader/crew', 'crew', 'lc', 'loader / crew'],
        'warehouse operative': ['warehouse', 'warehouse operator', 'picker packer', 'material handler'],
        'warehouse operator': ['warehouse', 'warehouse operative', 'picker packer', 'material handler'],
        'material handler': ['warehouse operative', 'warehouse operator', 'warehouse'],
        'industrial general labor': ['industrial general laborer', 'general labor'],
        'dishwasher': ['dishwash', 'dishwhasher'],
        'server': ['event server', 'buffet server', 'banquet server'],
        'event server': ['server', 'event servers'],
        'buffet server': ['server'],
        'barista': ['barista'],
        'bartender': ['bartender'],
        'cashier': ['cashier'],
        'assembler': ['assembler'],
        'repair technician': ['repair tech'],
        'concessions': ['concession stand worker', 'concession'],
        'forklift': ['forklift driver', 'forklift operator'],
        'forklift driver': ['forklift'],
        'driver': ['driver', 'cdl'],
    }
    aliases = role_aliases.get(a, [])
    for alias in aliases:
        if alias in b or b in alias:
            return True
    return False


def _fuzzy_client_match(a, b):
    for suffix in [', inc', ' inc', ', llc', ' llc', ', ltd', ' ltd']:
        a = a.replace(suffix, '')
        b = b.replace(suffix, '')
    a = a.strip().rstrip('.')
    b = b.strip().rstrip('.')
    if a == b:
        return True
    if a in b or b in a:
        return True
    a_words = a.split()
    b_words = b.split()
    if a_words and b_words and a_words[0] == b_words[0] and len(a_words[0]) > 3:
        return True
    return False


def _location_match(loc_a, loc_b):
    if not loc_a or not loc_b:
        return not loc_a and not loc_b
    a = loc_a.lower().strip()
    b = loc_b.lower().strip()
    if a == b:
        return True
    if a in b or b in a:
        return True
    if "/" in a:
        parts = a.split("/")
        return any(p.strip() == b for p in parts)
    if "/" in b:
        parts = b.split("/")
        return any(p.strip() == a for p in parts)
    return False


# =============================================================================
# HTML GENERATION
# =============================================================================

LOCATION_TO_STATE = {
    'chicago': 'IL', 'bedford park': 'IL', 'hodgkins': 'IL', 'joliet': 'IL', 'libertyville': 'IL',
    'dallas': 'TX', 'fort worth': 'TX', 'lancaster': 'TX', 'haslet': 'TX', 'flower mound': 'TX',
    'plano': 'TX', 'arlington': 'TX', 'austin': 'TX', 'houston': 'TX', 'san antonio': 'TX',
    'nashville': 'TN', 'lavergne': 'TN', 'la vergne': 'TN', 'lebanon': 'TN',
    'las vegas': 'NV', 'reno': 'NV', 'mccarran': 'NV', 'sparks': 'NV',
    'atlanta': 'GA', 'hapeville': 'GA', 'cartersville': 'GA',
    'orlando': 'FL', 'kissimmee': 'FL',
    'columbus': 'OH', 'grove city': 'OH', 'lockbourne': 'OH', 'cincinnati': 'OH',
    'middleburg heights': 'OH', 'west chester': 'OH', 'hebron': 'KY', 'erlanger': 'KY',
    'charlotte': 'NC', 'fort mill': 'SC',
    'phoenix': 'AZ',
    'washington, d.c.': 'DC', 'washington': 'DC',
    'logan township': 'NJ', 'paulsboro': 'NJ', 'south brunswick': 'NJ', 'newark': 'NJ',
    'philadelphia': 'PA',
    'inland empire': 'CA',
}


def _get_state_from_location(location):
    if not location:
        return 'Unknown'
    loc = location.lower().strip()
    for key, state in LOCATION_TO_STATE.items():
        if key in loc:
            return state
    m = re.search(r',\s*([A-Z]{2})\b', location)
    if m:
        return m.group(1)
    return 'Other'


def generate_html(processed_rows):
    """Generate the full HTML report."""
    # Sort by submission date ascending (chronological)
    processed_rows.sort(key=lambda r: r.get('submission_date') or date.max)

    all_states = set()
    all_owners = set()

    # Calculate KPIs -- only Live rows
    live_rows = [r for r in processed_rows if not r.get('is_declined')]
    declined_rows = [r for r in processed_rows if r.get('is_declined')]
    total_live = len(live_rows)
    total_declined = len(declined_rows)
    total_hc = sum(r['hc'] for r in live_rows if isinstance(r['hc'], int))
    high_priority = sum(1 for r in live_rows if isinstance(r['hc'], int) and r['hc'] >= 20)
    no_fhs = sum(1 for r in live_rows if r['fhs']['count'] == 0)
    no_indeed = sum(1 for r in live_rows if r['indeed']['count'] == 0)
    shifts_tbd = sum(1 for r in live_rows if not r['shifts'])

    rows_html = ""
    for r in processed_rows:
        is_declined = r.get('is_declined', False)

        if is_declined:
            bg = "#fef2f2"  # Light red for declined
        else:
            bg = "#ffffff"

        # Status badge: O = green for Live, D = red for Declined
        if is_declined:
            status_cell = '<td style="text-align:center;font-weight:bold;font-size:15px;color:#fff;background:#dc2626;padding:6px 10px;">D</td>'
        else:
            status_cell = '<td style="text-align:center;font-weight:bold;font-size:15px;color:#fff;background:#16a34a;padding:6px 10px;">O</td>'

        hc_display = str(r['hc']) if isinstance(r['hc'], int) else r['hc']
        shifts_tag = ('<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:4px;font-size:12px;">Yes</span>'
                      if r['shifts'] else
                      '<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:4px;font-size:12px;">No</span>')

        # Client display
        client_safe = r['client'].replace('&', '&amp;').replace('<', '&lt;')
        location_safe = r['location'].replace('&', '&amp;').replace('<', '&lt;') if r['location'] else ''
        role_safe = r['role'].replace('&', '&amp;').replace('<', '&lt;') if r['role'] else ''

        # Start date
        start_display = r['start_date'] if r['start_date'] else '&mdash;'

        # State for filter
        row_state = _get_state_from_location(r['location'])
        all_states.add(row_state)
        row_owner = r.get('owner', '').strip()
        if row_owner:
            all_owners.add(row_owner)

        row_status_class = "declined" if is_declined else "live"

        # Build metric cells
        if is_declined:
            # Declined: ALL metric columns blank (dashes)
            metric_cells = '''
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:right;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="text-align:center;color:#d1d5db;">&mdash;</td>
        <td style="color:#d1d5db;">&mdash;</td>'''
        else:
            fhs_open = f'{r["fhs"]["count"]}'
            fhs_rsvps = f'{r["fhs"]["rsvps"]:,}'

            # Indeed status emoji
            if r['indeed']['count'] > 0 and r['indeed']['spend'] > 0:
                indeed_st = "&#x2705;"  # checkmark
            elif r['indeed']['count'] > 0:
                indeed_st = "&#x26A0;&#xFE0F;"  # warning
            else:
                indeed_st = "&#x274C;"  # X mark
            indeed_camps = f'{r["indeed"]["count"]}'
            indeed_spend = f'${r["indeed"]["spend"]:,.0f}' if r['indeed']['spend'] > 0 else "$0"

            ob_created = f'{r["ob"]["created"]}'
            ob_verified = f'{r["ob"]["verified"]}'
            ob_rtb = f'{r["ob"]["rtb"]}'

            # Target = HC x 10
            hc_val = r['hc'] if isinstance(r['hc'], int) else 0
            target = hc_val * 10
            target_html = f'{target:,}' if target > 0 else '&mdash;'

            # Fill% = RTB / HC
            ob_rtb_val = r['ob']['rtb']
            if hc_val > 0:
                fill_pct = min((ob_rtb_val / hc_val) * 100, 999)
                fill_display = min(fill_pct, 100)
                fill_color = "#ef4444" if fill_pct < 25 else "#f59e0b" if fill_pct < 50 else "#22c55e"
                fill_html = f'''<div style="display:flex;align-items:center;gap:6px;">
            <div style="flex:1;background:#e5e7eb;border-radius:4px;height:14px;min-width:60px;">
                <div style="width:{fill_display:.1f}%;background:{fill_color};height:100%;border-radius:4px;"></div>
            </div>
            <span style="font-size:13px;font-weight:bold;min-width:40px;text-align:right;">{fill_pct:.0f}%</span>
        </div>'''
            else:
                fill_html = '<span style="color:#9ca3af;">&mdash;</span>'

            metric_cells = f'''
        <td style="text-align:center;font-weight:bold;font-size:15px;">{fhs_open}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{fhs_rsvps}</td>
        <td style="text-align:center;">{indeed_st}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{indeed_camps}</td>
        <td style="text-align:right;font-weight:bold;font-size:15px;">{indeed_spend}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_created}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_verified}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_rtb}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{target_html}</td>
        <td style="min-width:120px;">{fill_html}</td>'''

        rows_html += f'''<tr style="background:{bg};" data-state="{row_state}" data-owner="{row_owner}" data-status="{row_status_class}">
        {status_cell}
        <td><strong>{client_safe}</strong><br><span style="color:#6b7280;font-size:12px;">{location_safe}</span></td>
        <td>{row_owner}</td>
        <td>{role_safe}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{hc_display}</td>
        <td style="text-align:center;font-size:13px;">{start_display}</td>
        <td style="text-align:center;">{shifts_tag}</td>{metric_cells}
    </tr>
'''

    # Build filter options
    state_options = ''.join(f'<option value="{s}">{s}</option>' for s in sorted(all_states))
    owner_options = ''.join(f'<option value="{o}">{o}</option>' for o in sorted(all_owners))

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Revenue Requests Cross-Reference Dashboard &mdash; {REPORT_DATE}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f3f4f6; font-size: 14px; color: #111827; }}
        .header {{ background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%); color: white; padding: 24px 32px; }}
        .header h1 {{ font-size: 22px; font-weight: 700; }}
        .header p {{ font-size: 13px; opacity: 0.85; margin-top: 4px; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(6, 1fr); gap: 16px; padding: 20px 32px; }}
        .kpi-card {{ background: white; border-radius: 12px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); text-align: center; }}
        .kpi-card .number {{ font-size: 32px; font-weight: 800; line-height: 1.2; }}
        .kpi-card .label {{ font-size: 12px; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 4px; }}
        .kpi-card.red .number {{ color: #ef4444; }}
        .kpi-card.amber .number {{ color: #f59e0b; }}
        .kpi-card.blue .number {{ color: #2563eb; }}
        .kpi-card.green .number {{ color: #22c55e; }}
        .filters {{ display: flex; gap: 12px; padding: 0 32px 16px; align-items: center; flex-wrap: wrap; }}
        .filters select, .filters button {{ padding: 6px 12px; border-radius: 6px; border: 1px solid #d1d5db; font-size: 13px; background: white; cursor: pointer; }}
        .filters button {{ background: #2563eb; color: white; border: none; font-weight: 600; }}
        .filters button:hover {{ background: #1d4ed8; }}
        .table-container {{ margin: 0 32px 32px; background: white; border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow: hidden; }}
        .table-scroll {{ overflow-y: auto; max-height: calc(100vh - 60px); }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
        thead th {{ position: sticky; top: 0; z-index: 20; background: #1e3a5f; color: white; padding: 8px 10px; font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.3px; white-space: nowrap; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #e5e7eb; vertical-align: middle; font-size: 14px; }}
        tr:hover {{ background: #f0f9ff !important; }}
        .legend {{ display: flex; gap: 20px; padding: 8px 32px 0; font-size: 12px; color: #6b7280; }}
        .legend-item {{ display: flex; align-items: center; gap: 4px; }}
        .legend-dot {{ width: 14px; height: 14px; border-radius: 3px; display: inline-block; text-align: center; color: white; font-size: 10px; font-weight: bold; line-height: 14px; }}
        .count-badge {{ display: inline-block; background: #e5e7eb; color: #374151; font-size: 12px; font-weight: 600; padding: 2px 8px; border-radius: 10px; margin-left: 8px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Revenue Requests Cross-Reference Dashboard</h1>
        <p>Report 5 &mdash; Generated {REPORT_DATE} | Sorted by Submission Date (Ascending)</p>
    </div>

    <div class="kpi-grid">
        <div class="kpi-card green">
            <div class="number">{total_live}</div>
            <div class="label">Live Requests</div>
        </div>
        <div class="kpi-card blue">
            <div class="number">{total_hc:,}</div>
            <div class="label">Total HC</div>
        </div>
        <div class="kpi-card amber">
            <div class="number">{high_priority}</div>
            <div class="label">High Priority (HC&ge;20)</div>
        </div>
        <div class="kpi-card red">
            <div class="number">{no_fhs}</div>
            <div class="label">No FHS</div>
        </div>
        <div class="kpi-card red">
            <div class="number">{no_indeed}</div>
            <div class="label">No Indeed</div>
        </div>
        <div class="kpi-card amber">
            <div class="number">{shifts_tbd}</div>
            <div class="label">Shifts TBD</div>
        </div>
    </div>

    <div class="legend">
        <div class="legend-item"><span class="legend-dot" style="background:#16a34a;">O</span> Live</div>
        <div class="legend-item"><span class="legend-dot" style="background:#dc2626;">D</span> Declined</div>
        <div class="legend-item">&#x2705; Indeed Active &amp; Spending</div>
        <div class="legend-item">&#x26A0;&#xFE0F; Indeed Active, $0 Spend</div>
        <div class="legend-item">&#x274C; No Indeed</div>
    </div>

    <div class="filters">
        <select id="filterState" onchange="applyFilters()">
            <option value="">All States</option>
            {state_options}
        </select>
        <select id="filterOwner" onchange="applyFilters()">
            <option value="">All Owners</option>
            {owner_options}
        </select>
        <select id="filterStatus" onchange="applyFilters()">
            <option value="">All Statuses</option>
            <option value="live">Live</option>
            <option value="declined">Declined</option>
        </select>
        <button onclick="resetFilters()">Reset</button>
        <span id="rowCount" class="count-badge">{len(processed_rows)} rows</span>
    </div>

    <div class="table-container">
        <div class="table-scroll">
            <table>
                <thead>
                    <tr>
                        <th>St</th>
                        <th>Client</th>
                        <th>Owner</th>
                        <th>Role</th>
                        <th>HC</th>
                        <th>Start</th>
                        <th>Shifts</th>
                        <th colspan="2" style="background:#1e5631;text-align:center;">FHS</th>
                        <th colspan="3" style="background:#7c2d12;text-align:center;">Indeed</th>
                        <th colspan="3" style="background:#4338ca;text-align:center;">OB Funnel</th>
                        <th colspan="2" style="background:#92400e;text-align:center;">Fill</th>
                    </tr>
                    <tr>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;"></th>
                        <th style="top:34px;background:#1e5631;">Open</th>
                        <th style="top:34px;background:#1e5631;">RSVPs</th>
                        <th style="top:34px;background:#7c2d12;">St</th>
                        <th style="top:34px;background:#7c2d12;">Camps</th>
                        <th style="top:34px;background:#7c2d12;">Spend</th>
                        <th style="top:34px;background:#4338ca;">Created</th>
                        <th style="top:34px;background:#4338ca;">Verified</th>
                        <th style="top:34px;background:#4338ca;">RTB</th>
                        <th style="top:34px;background:#92400e;">Target</th>
                        <th style="top:34px;background:#92400e;">Fill%</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>
    </div>

    <script>
    function applyFilters() {{
        const state = document.getElementById('filterState').value;
        const owner = document.getElementById('filterOwner').value;
        const status = document.getElementById('filterStatus').value;
        const rows = document.querySelectorAll('tbody tr');
        let visible = 0;
        rows.forEach(row => {{
            let show = true;
            if (state && row.dataset.state !== state) show = false;
            if (owner && row.dataset.owner !== owner) show = false;
            if (status && row.dataset.status !== status) show = false;
            row.style.display = show ? '' : 'none';
            if (show) visible++;
        }});
        document.getElementById('rowCount').textContent = visible + ' rows';
    }}
    function resetFilters() {{
        document.getElementById('filterState').value = '';
        document.getElementById('filterOwner').value = '';
        document.getElementById('filterStatus').value = '';
        applyFilters();
    }}
    </script>
</body>
</html>'''

    return html


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("Loading Revenue Requests...")
    rev_rows = load_revenue_requests()
    print(f"  Loaded {len(rev_rows)} qualifying rows (Live + Declined)")

    print("Loading FHS Requisitions...")
    fhs_rows = load_fhs_requisitions()
    print(f"  Loaded {len(fhs_rows)} requisitions")

    print("Loading Indeed Campaigns (JobsCampaigns)...")
    indeed_rows = load_indeed_campaigns()
    print(f"  Loaded {len(indeed_rows)} campaigns")

    print("Loading OB Funnel...")
    ob_data = load_ob_funnel()
    print(f"  Loaded {len(ob_data)} client-location combos")

    print("\nCross-referencing data...")
    processed = []
    for r in rev_rows:
        if r['is_declined']:
            # Declined rows: no metric lookup
            r['fhs'] = {'count': 0, 'rsvps': 0}
            r['indeed'] = {'count': 0, 'spend': 0.0}
            r['ob'] = {'created': 0, 'verified': 0, 'rtb': 0}
        else:
            r['fhs'] = match_fhs(r['client'], r['location'], fhs_rows,
                                  submission_date=r.get('submission_date'), role=r.get('role'))
            r['indeed'] = match_indeed(r['client'], r['location'], indeed_rows, role=r.get('role'))
            r['ob'] = match_ob_funnel(r['client'], r['location'], ob_data, role=r.get('role'))
        processed.append(r)

    print(f"  Processed {len(processed)} rows")

    # ROW INTEGRITY CHECK
    expected = len(rev_rows)
    actual = len(processed)
    if actual != expected:
        print(f"  WARNING: Row count mismatch! Expected {expected}, got {actual}")
    else:
        print(f"  Row integrity OK: {actual} rows")

    print("\nGenerating HTML...")
    html = generate_html(processed)

    os.makedirs(os.path.dirname(OUTPUT_HTML), exist_ok=True)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Saved to {OUTPUT_HTML}")

    # Summary stats
    live_rows = [r for r in processed if not r.get('is_declined')]
    declined_rows = [r for r in processed if r.get('is_declined')]
    print(f"\n=== SUMMARY ===")
    print(f"  Live: {len(live_rows)}")
    print(f"  Declined: {len(declined_rows)}")
    print(f"  Total rows: {len(processed)}")
    total_hc = sum(r['hc'] for r in live_rows if isinstance(r['hc'], int))
    print(f"  Total HC (Live): {total_hc:,}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
