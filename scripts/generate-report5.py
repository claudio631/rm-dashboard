#!/usr/bin/env python3
"""
Report 5: Recruitment Request Dashboard
Generates an HTML dashboard cross-referencing Revenue Requests with FHS, Indeed, and OB Funnel data.

CRITICAL: Every "Live" row in Revenue Requests = one row in output. NO merging/dedup.
"""

import csv
import json
import re
import os
import sys
from collections import defaultdict
from datetime import datetime, date

import glob
import openpyxl

# =============================================================================
# FILE PATHS — auto-detect latest files
# =============================================================================
def _latest(pattern):
    files = sorted(glob.glob(os.path.expanduser(pattern)), key=os.path.getmtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No file found matching: {pattern}")
    return files[0]

REVENUE_CSV = _latest("~/Downloads/US_Recruitment_Requests__us_*.csv")
FHS_CSV = _latest("~/Downloads/requisitions-*.csv")
INDEED_CSV = _latest("~/Downloads/CampaignReport_Advanced_*.csv")
OB_FUNNEL_XLSX = _latest("~/Downloads/OB Funnel Custom Viewer*.xlsx")
REPORT_DATE = datetime.now().strftime("%Y-%m-%d")
OUTPUT_HTML = os.path.expanduser(f"~/RM-Team-Ai/docs/reports/recruitment-request-dashboard-{REPORT_DATE}.html")
OUTPUT_XLSX = os.path.expanduser(f"~/RM-Team-Ai/docs/reports/recruitment-request-dashboard-{REPORT_DATE}.xlsx")
FROZEN_DATA_PATH = os.path.expanduser("~/RM-Team-Ai/src/data/frozen-requests.json")

# =============================================================================
# CLIENT NAME NORMALIZATION MAPS
# =============================================================================
CLIENT_NAME_MAP = {
    "compass": ["culinaire", "culinaire international"],
    "dc flex": ["indeed flex", "indeed flex applications"],
    "food glorious": ["culinaire"],
    "bon appetit": ["bon appetit management company, inc", "bon appetit"],
    "legends": ["legends hospitality"],
    "levy": ["levy restaurants"],
    "btx": ["btx global logistics"],
    "lettuce": ["lettuce entertain you enterprises, inc"],
    "soho house": ["soho house austin"],
    "austin club": ["austin catering"],
    "vestals": ["vestals catering"],
    "rhino": ["rhino staging"],
    "power stop": ["power stop"],
    "uplift": ["uplift desk"],
    "ryerson": ["ryerson"],
    "solaren": ["solaren risk management"],
    "hei hotels": ["merritt hospitality llc"],
    "hyatt place": ["merritt hospitality llc", "hyatt", "hei"],
    "foxconn": ["foxconn"],
    "ctdi": ["ctdi"],
    "ingram": ["ingram content group"],
    "stord": ["stord, inc"],
    "ontrac": ["ontrac final mile"],
    "tennant": ["tennant solutions"],
    "continental battery": ["continental battery systems, inc."],
    "bath concepts": ["bath concepts"],
    "cort": ["cort"],
    "san antonio & austin": ["cort"],
    "assurant": ["assurant, inc (hyla)", "assurant"],
    "sxsw": ["sxsw"],
    "culinaire": ["culinaire", "culinaire international"],
    "material handler": ["ctdi"],  # "Material Handler" row is CTDI
}

# Metro/Location aliases - map suburb/alias to metro
LOCATION_ALIASES = {
    "dfw": "dallas",
    "vegas": "las vegas",
    "bedford park": "chicago",
    "hodgkins": "chicago",
    "fort worth": "dallas",
    "lancaster": "dallas",
    "flower mound": "dallas",
    "haslet": "dallas",
    "lavergne": "nashville",
    "la vergne": "nashville",
    "libertyville": "chicago",
    "grove city": "columbus",
    "mccarran": "las vegas",
    "middleburg heights": "columbus",
    "middle heights": "columbus",
    "south brunswick": "newark",
    "lebanon": "nashville",
    "lockbourne": "columbus",
    "logan township": "new jersey",
    "paulsboro": "new jersey",
    "hebron": "cincinnati",
    "west chester": "cincinnati",
    "washington": "washington, d.c.",
    "washington dc": "washington, d.c.",
    "washington, dc": "washington, d.c.",
    "washington d.c.": "washington, d.c.",
    "washington, d.c.": "washington, d.c.",
    "washington, d.c": "washington, d.c.",
    "bedford": "chicago",
    "arlington": "dallas",
    "kissimmee": "orlando",
    "desoto": "dallas",
    "plano": "dallas",
    "garland": "dallas",
    "mesquite": "dallas",
    "fairfield": "cincinnati",
    "cartersville": "atlanta",
    "erlanger": "cincinnati",
    "fort mill": "charlotte",
    "nw atlanta": "atlanta",
    "wimberley": "austin",
    "pasadena": "houston",
    "inland empire": "inland empire",
    "reno": "reno",
    "joliet": "chicago",
    "sparks": "reno",
    "hapeville": "atlanta",
}


def normalize_location(loc_str):
    """Normalize a location string to a canonical metro name."""
    if not loc_str:
        return ""
    loc = loc_str.lower().strip()

    # Check aliases FIRST on the raw string (before any stripping)
    for alias, metro in LOCATION_ALIASES.items():
        if alias in loc:
            return metro

    # Check common metros on raw string
    for metro in ["chicago", "dallas", "austin", "nashville", "las vegas",
                  "orlando", "houston", "atlanta", "columbus", "cincinnati",
                  "charlotte", "phoenix", "reno", "newark", "new jersey",
                  "philadelphia", "cleveland", "san antonio", "inland empire",
                  "washington, d.c."]:
        if metro in loc:
            return metro

    # Remove state abbreviations and zip codes (pattern: ", XX" or ", XX 12345")
    loc = re.sub(r',\s*[A-Za-z]{2}\s*\d{0,5}\s*$', '', loc_str.lower().strip())
    loc = loc.strip().rstrip(',').strip()

    # Re-check aliases after cleaning
    for alias, metro in LOCATION_ALIASES.items():
        if alias in loc:
            return metro

    # Re-check metros after cleaning
    for metro in ["chicago", "dallas", "austin", "nashville", "las vegas",
                  "orlando", "houston", "atlanta", "columbus", "cincinnati",
                  "charlotte", "phoenix", "reno", "newark", "new jersey",
                  "philadelphia", "cleveland", "san antonio", "inland empire",
                  "washington, d.c."]:
        if metro in loc:
            return metro

    return loc


def normalize_client_for_match(client_short):
    """Get list of possible FHS/Indeed/Funnel client names for a revenue request client."""
    c = client_short.lower().strip()
    # Direct map lookup
    for key, vals in CLIENT_NAME_MAP.items():
        if key == c or c.startswith(key):
            return vals
    # Return as-is
    return [c]


def extract_client_and_location(client_job_str):
    """Extract short client name and location from the Client-Job column."""
    s = client_job_str.strip()

    # Special cases - exact or prefix matches on known entries
    sl = s.lower()
    if sl.startswith("need forklift"):
        return "Stord", "Las Vegas"
    if "aba nashville" in sl or (sl.startswith("aba") and "lettuce" in sl):
        return "Lettuce", "Nashville"
    if "texas motor speedway" in sl:
        return "Levy", "Dallas"
    if sl == "material handler":
        return "CTDI", "Dallas"
    if sl.startswith("san antonio & austin"):
        return "CORT", "San Antonio/Austin"
    if sl.startswith("assurant"):
        return "Assurant", "Nashville"
    if sl == "foxconn":
        return "Foxconn", "Dallas"  # Foxconn DFW (Fort Worth)
    if sl == "rhino staging":
        return "Rhino", "Austin"  # Rhino Staging is Austin-based
    if sl.startswith("sxsw"):
        return "SXSW", "Austin"
    if "vestals catering" in sl:
        return "Vestals", "Austin"  # Vestals Catering is Austin-based
    if "obama foundation" in sl:
        return "Bon Appetit", "Chicago"  # Obama Foundation is in Chicago
    if "dos equis pavilion" in sl and "legends" in sl:
        if "at&t" not in sl:
            return "Legends", "Dallas"  # Dos Equis Pavilion is in Dallas
    if "at&t stadium" in sl and "legends" in sl:
        return "Legends", "Dallas"  # AT&T Stadium is in Dallas
    if "soldier field" in sl and "levy" in sl:
        return "Levy", "Chicago"  # Soldier Field is in Chicago

    # Try pattern: "Client - Location - Role" or "Client - Location stuff"
    # First token before first dash is usually the client
    parts = re.split(r'\s*-\s*', s, maxsplit=3)

    client = parts[0].strip()
    # Remove leading parens like "(Culinaire)"
    if client.startswith("("):
        m = re.match(r'\(([^)]+)\)\s*(.*)', client)
        if m:
            client = m.group(2).strip() if m.group(2).strip() else m.group(1).strip()

    location = ""
    if len(parts) >= 2:
        loc_part = parts[1].strip()
        # Check if second part looks like a location
        # Locations often have city names, states, addresses
        location = loc_part

    # Try to extract city from the location part
    loc = _extract_city(location, s)

    # Special client cleanups
    if client.upper() == "CORT" or client.lower() == "cort":
        client = "CORT"
    elif "culinaire" in client.lower():
        client = "Culinaire"
    elif "food glorious" in client.lower():
        client = "Culinaire"
    elif "compass" in client.lower():
        client = "Compass"
    elif "bon appetit" in client.lower():
        client = "Bon Appetit"
    elif "dc flex" in client.lower():
        client = "DC Flex"
    elif "foxconn" in client.lower():
        client = "Foxconn"
    elif "power stop" in client.lower():
        client = "Power Stop"
    elif "legends" in client.lower():
        client = "Legends"
    elif "levy" in client.lower():
        client = "Levy"
    elif "ingram" in client.lower():
        client = "Ingram"
    elif "ctdi" in client.lower():
        client = "CTDI"
    elif "bath concepts" in client.lower():
        client = "Bath Concepts"
    elif "btx" in client.lower():
        client = "BTX"
    elif "solaren" in client.lower():
        client = "Solaren"
    elif "hei hotels" in client.lower():
        client = "HEI Hotels"
    elif "hyatt" in client.lower():
        client = "Hyatt Place"
    elif "rhino" in client.lower():
        client = "Rhino"
    elif "uplift" in client.lower():
        client = "Uplift"
    elif "ryerson" in client.lower():
        client = "Ryerson"
    elif "lettuce" in client.lower():
        client = "Lettuce"
    elif "soho" in client.lower():
        client = "Soho House"
    elif "austin club" in client.lower():
        client = "Austin Club"
    elif "vestals" in client.lower():
        client = "Vestals"
    elif "sxsw" in client.lower():
        client = "SXSW"

    return client, loc


def _extract_city(loc_part, full_str):
    """Try to extract a city name from a location part."""
    # Check for known cities in the string
    cities = {
        "chicago": "Chicago",
        "dallas": "Dallas",
        "dfw": "Dallas",
        "austin": "Austin",
        "nashville": "Nashville",
        "las vegas": "Las Vegas",
        "vegas": "Las Vegas",
        "orlando": "Orlando",
        "houston": "Houston",
        "atlanta": "Atlanta",
        "columbus": "Columbus",
        "cincinnati": "Cincinnati",
        "charlotte": "Charlotte",
        "phoenix": "Phoenix",
        "reno": "Reno",
        "san antonio": "San Antonio",
        "washington": "Washington, D.C.",
        "fort worth": "Dallas",
        "bedford park": "Chicago",
        "hodgkins": "Chicago",
        "la vergne": "Nashville",
        "lavergne": "Nashville",
        "grove city": "Columbus",
        "libertyville": "Chicago",
        "bedford": "Chicago",
    }

    check_str = full_str.lower()
    for key, city in cities.items():
        if key in check_str:
            return city

    # Fallback: return the loc_part cleaned up
    if loc_part:
        # Remove address-like parts
        cleaned = re.sub(r'\d+\s+\w+\s+(dr|drive|st|street|ave|blvd|rd|way)\b.*', '', loc_part, flags=re.IGNORECASE).strip()
        if cleaned:
            return cleaned
    return loc_part


# Source: src/data/clients-and-markets.xlsx
MARKET_OWNERS = {
    "claudio": [
        "austin", "houston", "charlotte", "fort mill", "orlando",
        "las vegas", "reno", "washington, d.c.", "washington", "monroe", "phoenix",
        "logan township"
    ],
    "craig": [
        "columbus", "cincinnati", "hebron", "chicago", "atlanta",
        "dallas", "north haven", "south brunswick",
        "hamilton", "kearny", "nashville", "middleburg heights",
        "erlanger", "newark", "paulsboro", "mccarran"
    ],
}


def get_owner_by_market(location):
    """Determine owner based on market assignment list, not revenue request email."""
    loc_norm = normalize_location(location)
    for owner, markets in MARKET_OWNERS.items():
        for market in markets:
            if market in loc_norm or loc_norm in market:
                return owner.title()
    return ""


def parse_owner(email):
    """Extract first name from email (fallback only)."""
    owners = {
        "claudio.santos": "Claudio",
        "craig.freeman": "Craig",
        "megan.mccue": "Megan",
        "angela": "Angela",
        "lacey.henderson": "Lacey",
        "david.starkman": "David",
    }
    email_lower = email.lower().strip()
    for key, name in owners.items():
        if key in email_lower:
            return name
    m = re.match(r'([a-z]+)', email_lower)
    return m.group(1).title() if m else email


def parse_headcount(hc_str):
    """Parse headcount from string. Return int or string 'LFDTW'."""
    s = hc_str.strip()
    if not s:
        return 0
    if s.upper() == "LFDTW":
        return "LFDTW"
    # Try to parse number
    m = re.search(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 0


def has_shifts(shift_info):
    """Determine if shifts are posted (Yes) or not (No/TBD)."""
    if not shift_info:
        return False
    s = shift_info.lower().strip()
    # Indicators of no shifts
    no_indicators = ["not posted", "tbd", "no posted", "shifts not posted",
                     "need to immediately start recruiting", "not yet",
                     "will be based on client posts"]
    for ind in no_indicators:
        if ind in s:
            return False
    # If there are time patterns, shifts exist
    if re.search(r'\d+\s*(am|pm|a\.m|p\.m)', s, re.IGNORECASE):
        return True
    if re.search(r'(monday|tuesday|wednesday|thursday|friday|saturday|sunday|mon|tue|wed|thu|fri|sat|sun|m-f)', s, re.IGNORECASE):
        return True
    # Default: if there's substantial content, assume yes
    if len(s) > 10:
        return True
    return False


def parse_currency(val_str):
    """Parse a currency string like '$1,234.56' to float."""
    if not val_str or val_str.strip() == '-':
        return 0.0
    s = val_str.strip().replace('$', '').replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# =============================================================================
# LOAD DATA
# =============================================================================

def load_revenue_requests():
    """Load all Live and Declined rows from revenue requests CSV."""
    rows = []
    with open(REVENUE_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 12:
                continue
            status = row[2].strip()
            if "Live" not in status and "Declined" not in status and "Complete" not in status:
                continue
            is_declined = "Declined" in status
            is_complete = "Complete" in status

            client_job = row[0].strip()
            owner_email = row[1].strip()
            role_col = row[5].strip()
            other_role = row[6].strip() if len(row) > 6 else ""
            shift_info = row[9].strip() if len(row) > 9 else ""
            start_date = row[10].strip() if len(row) > 10 else ""
            hc_str = row[11].strip() if len(row) > 11 else "0"
            submission_str = row[17].strip() if len(row) > 17 else ""

            # Parse submission date (format: "25/03/2026, 12:55")
            submission_date = None
            if submission_str:
                try:
                    submission_date = datetime.strptime(submission_str.split(',')[0].strip(), '%d/%m/%Y').date()
                except ValueError:
                    pass

            client, location = extract_client_and_location(client_job)
            # Owner based on market assignment, fallback to email
            owner = get_owner_by_market(location) or parse_owner(owner_email)
            hc = parse_headcount(hc_str)

            # Role: use Role column, or other_role if Role is "Other", or extract from client_job
            role = role_col
            if role.lower() == "other" and other_role:
                role = other_role
            if not role:
                # Try to extract from client_job (last part after dashes)
                parts = re.split(r'\s*-\s*', client_job)
                if len(parts) >= 3:
                    role = parts[-1].strip()
                    # Clean up address-like roles
                    if re.match(r'^\d', role) or len(role) > 60:
                        role = parts[2].strip() if len(parts) > 2 else ""

            shifts = has_shifts(shift_info)

            # Format start date as mmm/dd/yyyy
            formatted_date = start_date
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    dt = datetime.strptime(start_date, fmt)
                    formatted_date = dt.strftime('%b/%d/%Y')
                    break
                except ValueError:
                    continue

            rows.append({
                'client_job': client_job,
                'client': client,
                'location': location,
                'owner': owner,
                'role': role,
                'hc': hc,
                'start_date': formatted_date,
                'shifts': shifts,
                'shift_info': shift_info,
                'is_declined': is_declined,
                'is_complete': is_complete,
                'submission_date': submission_date,
            })

    return rows


# =============================================================================
# FROZEN DATA — Snapshot Complete rows so they don't update on future runs
# =============================================================================

def _frozen_key(row):
    """Build a stable key for a revenue request row."""
    sub = row.get('submission_date')
    sub_str = sub.isoformat() if sub else ''
    return f"{row['client_job']}|{sub_str}"


def load_frozen_data():
    """Load frozen snapshots from disk."""
    if os.path.exists(FROZEN_DATA_PATH):
        with open(FROZEN_DATA_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_frozen_data(frozen):
    """Save frozen snapshots to disk."""
    os.makedirs(os.path.dirname(FROZEN_DATA_PATH), exist_ok=True)
    with open(FROZEN_DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(frozen, f, indent=2)


def load_fhs_requisitions():
    """Load FHS requisitions as individual rows for flexible filtering.
    Returns list of dicts with: client, location, loc_norm, status, rsvps, last_updated (date)."""
    rows = []
    with open(FHS_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 13:
                continue
            last_updated_str = row[0].strip()
            job_title = row[3].strip().lower() if len(row) > 3 else ""
            client = row[5].strip()
            location = row[6].strip()
            status = row[12].strip()
            rsvps = int(row[10]) if row[10].strip().isdigit() else 0

            # Parse date from last_updated (format: 2026-03-25T16:14:16...)
            try:
                req_date = datetime.strptime(last_updated_str[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                req_date = None

            loc_norm = normalize_location(location)
            rows.append({
                'client': client.lower().strip(),
                'loc_norm': loc_norm,
                'job_title': job_title,
                'status': status,
                'rsvps': rsvps,
                'date': req_date,
            })
    return rows


def load_indeed_campaigns():
    """Load Indeed campaigns from CampaignReport_Advanced CSV as individual rows.
    Supports two formats:
      - Legacy: adKey(0), Campaign(1), Source(2), Status(3), ..., Budget spent(9)
      - New:    Campaign(0), Source(1), Status(2), Total budget(3), Budget spent(4)
    Filter to ACTIVE campaigns only.
    """
    rows = []
    with open(INDEED_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            campaign_name = row.get('Campaign', '').strip()
            status = row.get('Status', '').strip()
            spend = parse_currency(row.get('Budget spent', row.get('Total cost', '0')))

            # Only ACTIVE campaigns
            if status != "ACTIVE":
                continue

            # Skip empty/header-like rows
            if not campaign_name or campaign_name.lower() == "priority campaign":
                continue

            # Parse campaign name to extract client, location, and role
            client, location, role = _parse_campaign_name_with_role(campaign_name)
            if not client:
                continue

            loc_norm = normalize_location(location)
            rows.append({
                'client': client.lower().strip(),
                'loc_norm': loc_norm,
                'role': role.lower().strip() if role else '',
                'spend': spend,
                'name': campaign_name,
            })

    return rows


def _parse_campaign_name(name):
    """Parse Indeed campaign name to extract client and location.
    Format variations:
    - US - B2C - {category} - {client} - {role} - {location} - {dates}
    - US-B2C-{category}-{client}-{role}-{location}-{dates}
    - US - B2C - Stord- Picker Packer - Atlanta, GA - ...
    - US - Industrial - ...
    """
    s = name.strip()

    # Split by " - " or "-" as separators
    parts = re.split(r'\s*-\s*', s)

    # Known client names that appear in campaigns
    known_clients = {
        'cort', 'johnstone', 'bon appetit', 'evergreen', 'ontrac', 'tennant',
        'legends', 'culinaire', 'afc', 'foxconn', 'stord', 'powerstop',
        'power stop', 'powerststop', 'ryerson', 'bath concepts', 'btx',
        'ctdi', 'ingram', 'indeed flex', 'indeed', 'lettuce',
        'continental battery', 'hyatt', 'ac', 'levy', 'solaren',
        'uplift', 'soho house', 'rhino', 'vestals', 'sxsw',
        'hei hotels', 'austin club'
    }

    client = None
    location = None

    # Strategy: find location first (pattern "City, ST"), then find client
    # by scanning parts between category and location for known clients
    loc_idx = None
    for i, part in enumerate(parts):
        p = part.strip()
        if re.match(r'^[A-Za-z\s\.]+,\s*[A-Z]{2}$', p):
            location = p
            loc_idx = i
            break

    if loc_idx is None:
        # No standard "City, ST" location found; try to identify client from all parts
        candidate_parts = []
        skip_words_ext = {'us', 'b2c', 'industrial', 'hospitality', 'hiring event'}
        for p in [x.strip() for x in parts]:
            if p.lower() in skip_words_ext:
                continue
            candidate_parts.append(p)

        combined = ' '.join(candidate_parts).lower()
        multi_word_clients2 = sorted(known_clients, key=len, reverse=True)
        for kc in multi_word_clients2:
            if kc in combined:
                client = kc
                break
        if client:
            # Check for known location abbreviations in the parts
            known_loc_abbrevs = {
                'dfw': 'Dallas, TX',
                'atx': 'Austin, TX',
                'chi': 'Chicago, IL',
            }
            for p in candidate_parts:
                pl = p.lower().strip()
                if pl in known_loc_abbrevs:
                    location = known_loc_abbrevs[pl]
                    break

            # If still no location, infer from client context
            if not location:
                cl_lower = client.lower()
                client_default_locations = {
                    'bath concepts': 'Libertyville, IL',
                    'foxconn': 'Fort Worth, TX',
                    'ryerson': 'Dallas, TX',
                }
                location = client_default_locations.get(cl_lower, '')
            return client, location if location else ""
        return None, None

    # Location was found; reconstruct the portion before location to find client
    before_loc = [p.strip() for p in parts[:loc_idx]]

    # Skip known prefix words
    skip_words = {'us', 'b2c', 'industrial', 'hospitality', 'hiring event'}

    # Find client by looking for known client names in the parts
    # Build a combined string of parts after skipping prefixes
    candidate_parts = []
    for p in before_loc:
        if p.lower() in skip_words:
            continue
        candidate_parts.append(p)

    # Try to match known clients from the candidate parts
    combined = ' '.join(candidate_parts).lower()

    # Try multi-word clients first
    multi_word_clients = sorted(known_clients, key=len, reverse=True)
    for kc in multi_word_clients:
        if kc in combined:
            client = kc
            break

    # If not found via known list, first candidate part is likely the client
    if not client and candidate_parts:
        client = candidate_parts[0]

    if not client:
        return None, None

    # Normalize some campaign client names
    cl = client.lower().strip()
    if cl in ('powerstop', 'powerststop'):
        client = 'Power Stop'
    elif cl == 'ac':
        client = 'Austin Club'  # AC = Austin Club

    return client, location if location else ""


def _parse_campaign_name_with_role(name):
    """Parse Indeed campaign name to extract client, location, AND role.
    Format: US - B2C - {category} - {client} - {role} - {location} - {dates}
    Returns (client, location, role).
    """
    s = name.strip()
    parts = re.split(r'\s*-\s*', s)

    known_clients = {
        'cort', 'johnstone', 'bon appetit', 'evergreen', 'ontrac', 'tennant',
        'legends', 'culinaire', 'afc', 'foxconn', 'stord', 'powerstop',
        'power stop', 'powerststop', 'ryerson', 'bath concepts', 'btx',
        'ctdi', 'ingram', 'indeed flex', 'indeed', 'lettuce',
        'continental battery', 'hyatt', 'ac', 'levy', 'solaren',
        'uplift', 'soho house', 'rhino', 'vestals', 'sxsw',
        'hei hotels', 'austin club'
    }

    skip_words = {'us', 'b2c', 'industrial', 'hospitality', 'hiring event'}

    client = None
    location = None
    role = None

    # Find location (pattern "City, ST")
    loc_idx = None
    for i, part in enumerate(parts):
        p = part.strip()
        if re.match(r'^[A-Za-z\s\.]+,\s*[A-Z]{2}$', p):
            location = p
            loc_idx = i
            break

    if loc_idx is not None:
        before_loc = [p.strip() for p in parts[:loc_idx]]
        candidate_parts = [p for p in before_loc if p.lower() not in skip_words]

        # Find client in candidates
        combined = ' '.join(candidate_parts).lower()
        client_idx_in_candidates = None
        for kc in sorted(known_clients, key=len, reverse=True):
            if kc in combined:
                client = kc
                # Find which candidate part contains the client
                for ci, cp in enumerate(candidate_parts):
                    if kc in cp.lower():
                        client_idx_in_candidates = ci
                        break
                break

        if not client and candidate_parts:
            client = candidate_parts[0]
            client_idx_in_candidates = 0

        # Role = parts between client and location
        if client_idx_in_candidates is not None and client_idx_in_candidates + 1 < len(candidate_parts):
            role_parts = candidate_parts[client_idx_in_candidates + 1:]
            if role_parts:
                role = ' '.join(role_parts).strip()
    else:
        # No "City, ST" location — extract client and role from parts
        candidate_parts = [p.strip() for p in parts if p.strip().lower() not in skip_words]
        # Filter out date-like parts (e.g., "March 26", "March 30, 2026")
        non_date_parts = []
        for cp in candidate_parts:
            if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d', cp, re.IGNORECASE):
                break  # Stop at first date part
            non_date_parts.append(cp)

        combined = ' '.join(non_date_parts).lower()
        client_idx_in_nondates = None
        for kc in sorted(known_clients, key=len, reverse=True):
            if kc in combined:
                client = kc
                for ci, cp in enumerate(non_date_parts):
                    if kc in cp.lower():
                        client_idx_in_nondates = ci
                        break
                break
        if not client and non_date_parts:
            client = non_date_parts[0]
            client_idx_in_nondates = 0

        # Role = parts after client, before dates
        if client_idx_in_nondates is not None and client_idx_in_nondates + 1 < len(non_date_parts):
            role_parts = non_date_parts[client_idx_in_nondates + 1:]
            if role_parts:
                role = ' '.join(role_parts).strip()

        # Infer location from client defaults
        if client and not location:
            client_default_locations = {
                'levy': 'Chicago, IL',
                'bath concepts': 'Libertyville, IL',
                'foxconn': 'Fort Worth, TX',
                'ryerson': 'Dallas, TX',
                'sxsw': 'Austin, TX',
            }
            location = client_default_locations.get(client.lower(), '')

    # Normalize client names
    if client:
        cl = client.lower().strip()
        if cl in ('powerstop', 'powerststop'):
            client = 'Power Stop'
        elif cl == 'ac':
            client = 'Austin Club'

    # Clean up role
    if role:
        role = re.sub(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s*\d+.*', '', role, flags=re.IGNORECASE).strip()
        role = re.sub(r'\b(premium|additional\s+location\w*)\b', '', role, flags=re.IGNORECASE).strip()
        role = role.strip(' -,')

    return client, location if location else "", role if role else ""


def load_ob_funnel():
    """Load OB Funnel data from Excel, grouped by client+location+role.
    Returns dict keyed by (client_lower, loc_norm, role_lower).
    'Total' rows are stored with role='__total__' for fallback matching.
    """
    data = defaultdict(lambda: {'created': 0, 'verified': 0, 'rtb': 0})

    wb = openpyxl.load_workbook(OB_FUNNEL_XLSX, data_only=True)
    ws = wb.active

    current_client = None
    current_loc = None
    current_role = None
    grand_total_col = ws.max_column  # Last column is Grand Total

    # Detect format: 2-col (old) or 3-col (new with role slice)
    header_row = [c.value for c in ws[1]]
    has_role_col = len(header_row) > 2 and header_row[2] and 'Slice' in str(header_row[2])
    stage_col = 3 if has_role_col else 2

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]

        if cells[0] is not None and str(cells[0]).strip():
            current_client = str(cells[0]).strip()
        if cells[1] is not None and str(cells[1]).strip():
            current_loc = str(cells[1]).strip()
        if has_role_col and cells[2] is not None and str(cells[2]).strip():
            current_role = str(cells[2]).strip()

        metric = cells[stage_col]
        if not metric or not current_client or not current_loc:
            continue

        metric_str = str(metric).strip()
        grand_total = cells[grand_total_col - 1]  # 0-indexed
        if grand_total is None:
            grand_total = 0
        try:
            val = int(float(grand_total))
        except (ValueError, TypeError):
            val = 0

        loc_norm = normalize_location(current_loc)
        role_key = '__total__' if (not has_role_col or not current_role or current_role.lower() == 'total') else current_role.lower().strip()
        key = (current_client.lower().strip(), loc_norm, role_key)

        if "Worker Accounts Created" in metric_str:
            data[key]['created'] += val
        elif "1st Role Verified" in metric_str:
            data[key]['verified'] += val
        elif "1st OB Task Completed" in metric_str:
            data[key]['rtb'] += val

    wb.close()
    return data


# =============================================================================
# CROSS-REFERENCE MATCHING
# =============================================================================

def match_fhs(client_short, location, fhs_rows, submission_date=None, role=None):
    """Find FHS data matching a revenue request client+location+role.
    - Open count: only open + auto-paused status reqs
    - Interview (RSVPs): ALL statuses, but only reqs created on/after submission_date
    - Role filtering: when a role is provided, only match FHS reqs with matching job_title
    """
    possible_clients = normalize_client_for_match(client_short)
    loc_norm = normalize_location(location)

    total = {'count': 0, 'rsvps': 0}

    for fhs_row in fhs_rows:
        # Check client+location match
        client_match = False
        for pc in possible_clients:
            pc_lower = pc.lower().strip()
            if (pc_lower in fhs_row['client'] or fhs_row['client'] in pc_lower or
                    _fuzzy_client_match(pc_lower, fhs_row['client'])):
                client_match = True
                break
        if not client_match:
            continue
        if not _location_match(loc_norm, fhs_row['loc_norm']):
            continue

        # Role filtering: if revenue request has a role, FHS req must match it
        if role:
            if not fhs_row['job_title']:
                continue
            if not _role_match(role, fhs_row['job_title']):
                continue

        # Open count: only open + auto-paused
        if fhs_row['status'] in ('open', 'auto-paused'):
            total['count'] += 1

        # Interview (RSVPs): all statuses, but only if created after submission date
        if submission_date and fhs_row['date']:
            if fhs_row['date'] >= submission_date:
                total['rsvps'] += fhs_row['rsvps']
        elif not submission_date:
            # No submission date available, count all
            total['rsvps'] += fhs_row['rsvps']

    # Fallback: if no client match found, try location + role only (any client)
    if total['count'] == 0 and total['rsvps'] == 0 and role and loc_norm:
        for fhs_row in fhs_rows:
            if not _location_match(loc_norm, fhs_row['loc_norm']):
                continue
            if not fhs_row['job_title'] or not _role_match(role, fhs_row['job_title']):
                continue
            if fhs_row['status'] in ('open', 'auto-paused'):
                total['count'] += 1
            if submission_date and fhs_row['date']:
                if fhs_row['date'] >= submission_date:
                    total['rsvps'] += fhs_row['rsvps']
            elif not submission_date:
                total['rsvps'] += fhs_row['rsvps']

    return total


def _role_match(rev_role, fhs_job_title):
    """Check if a revenue request role matches an FHS job title."""
    a = rev_role.lower().strip()
    b = fhs_job_title.lower().strip()
    if not a or not b:
        return True  # No role to filter on → match all
    if a == b:
        return True
    if a in b or b in a:
        return True
    # Common aliases
    role_aliases = {
        'food preps': ['prep cook', 'prep'],
        'prep cook': ['food preps', 'prep'],
        'prep': ['prep cook', 'food preps'],
        'loader/crew': ['loader', 'crew', 'lc'],
        'loader': ['loader/crew', 'crew', 'lc'],
        'lc': ['loader/crew', 'loader', 'crew'],
        'warehouse operative': ['warehouse', 'warehouse operator', 'picker packer', 'material handler', 'wo'],
        'warehouse operator': ['warehouse', 'warehouse operative', 'picker packer', 'material handler', 'wo'],
        'wo': ['warehouse operative', 'warehouse operator', 'warehouse'],
        'picker packer': ['warehouse operative', 'warehouse operator', 'warehouse', 'pp'],
        'pp': ['picker packer', 'warehouse operative', 'warehouse'],
        'material handler': ['warehouse operative', 'warehouse operator', 'warehouse'],
        'industrial general labor': ['industrial general laborer', 'general labor', 'general laborer', 'igl'],
        'industrial general laborer': ['industrial general labor', 'general labor', 'general laborer', 'igl'],
        'igl': ['industrial general labor', 'industrial general laborer', 'general labor'],
        'dishwasher': ['dishwash', 'dishwhasher'],
        'dishwash': ['dishwasher', 'dishwhasher'],
        'dishwhasher': ['dishwasher', 'dishwash'],
        'server': ['event server', 'buffet server', 'banquet server'],
        'event server': ['server'],
        'buffet server': ['server'],
        'barista': ['barista'],
        'bartender': ['bartender'],
        'cashier': ['cashier'],
        'assembler': ['assembler'],
        'repair technician': ['repair tech'],
        'repair tech': ['repair technician'],
        'htl': ['hospitality', 'hospitality general labor'],
        'hospitality': ['htl', 'hospitality general labor'],
        'concession stand worker': ['concessions', 'concession'],
        'concessions': ['concession stand worker'],
        'multi': ['hospitality', 'general'],
        'forklift': ['forklift driver', 'forklift operator'],
        'forklift driver': ['forklift'],
    }
    aliases = role_aliases.get(a, [])
    for alias in aliases:
        if alias in b or b in alias:
            return True
    return False


def match_indeed(client_short, location, indeed_rows, role=None):
    """Find Indeed campaign data matching a revenue request.
    Priority:
      1. client + location + role (exact match)
      2. location + role only (fallback — picks up Indeed Flex / other client campaigns for same role+metro)
    """
    possible_clients = normalize_client_for_match(client_short)
    loc_norm = normalize_location(location)

    total = {'count': 0, 'spend': 0.0}

    for ind_row in indeed_rows:
        # Check client match
        client_match = False
        for pc in possible_clients:
            pc_lower = pc.lower().strip()
            if (pc_lower in ind_row['client'] or ind_row['client'] in pc_lower or
                    _fuzzy_client_match(pc_lower, ind_row['client'])):
                client_match = True
                break
        if not client_match:
            continue
        if not _location_match(loc_norm, ind_row['loc_norm']):
            continue

        # Role filtering: if revenue request has a role, campaign must match it
        if role:
            if not ind_row['role']:
                continue  # Campaign has no role parsed — skip when we need a specific role
            if not _role_match(role, ind_row['role']):
                continue

        total['count'] += 1
        total['spend'] += ind_row['spend']

    # Fallback: if no client match found, try location + role only (any client)
    if total['count'] == 0 and role and loc_norm:
        for ind_row in indeed_rows:
            if not _location_match(loc_norm, ind_row['loc_norm']):
                continue
            if ind_row['role'] and _role_match(role, ind_row['role']):
                total['count'] += 1
                total['spend'] += ind_row['spend']

    return total


# OB Funnel tracks workers under the end-client, not "Indeed Flex Applications"
# DC Flex in Washington D.C.: Loader/Crew under CORT, hospitality roles under OnTrac Final Mile
OB_FUNNEL_CLIENT_OVERRIDES = {
    "dc flex": ["cort", "ontrac final mile", "ontrac"],
}


def match_ob_funnel(client_short, location, ob_data, role=None):
    """Find OB Funnel data matching a revenue request client+location+role.
    Priority order:
      1. client + location + role (exact role match)
      2. client + location + '__total__' (fallback to Total row)
      3. location-only + role (across all clients)
      4. location-only + '__total__' (across all clients)
    """
    possible_clients = normalize_client_for_match(client_short)
    override_clients = OB_FUNNEL_CLIENT_OVERRIDES.get(client_short.lower().strip(), [])
    if override_clients:
        possible_clients = possible_clients + override_clients
    loc_norm = normalize_location(location)

    total = {'created': 0, 'verified': 0, 'rtb': 0}
    matched_keys = set()

    def _try_match(check_client=True, use_role=True):
        """Try matching with given constraints. Returns True if matched."""
        found = False
        for (ob_client, ob_loc, ob_role), vals in ob_data.items():
            key = (ob_client, ob_loc, ob_role)
            if key in matched_keys:
                continue

            # Role filter
            if use_role and role:
                if not _role_match(role, ob_role):
                    continue
            elif use_role and not role:
                # No role specified — use __total__ only
                if ob_role != '__total__':
                    continue
            else:
                # Not using role — use __total__ only
                if ob_role != '__total__':
                    continue

            # Location filter
            if not _location_match(loc_norm, ob_loc):
                continue

            # Client filter
            if check_client:
                client_match = False
                for pc in possible_clients:
                    pc_lower = pc.lower().strip()
                    if (pc_lower in ob_client or ob_client in pc_lower or
                            _fuzzy_client_match(pc_lower, ob_client)):
                        client_match = True
                        break
                if not client_match:
                    continue

            matched_keys.add(key)
            total['created'] += vals['created']
            total['verified'] += vals['verified']
            total['rtb'] += vals['rtb']
            found = True
        return found

    # 1. Try client + location + role
    if role:
        if _try_match(check_client=True, use_role=True):
            return total

    # 2. Try client + location + __total__
    if _try_match(check_client=True, use_role=False):
        return total

    # No location-only fallback for OB Funnel — summing all clients in a city
    # would produce inflated numbers (e.g. all Chicago clients for a Compass request)
    return total


def _fuzzy_client_match(a, b):
    """Basic fuzzy matching for client names."""
    # Strip common suffixes
    for suffix in [', inc', ' inc', ', llc', ' llc', ', ltd', ' ltd']:
        a = a.replace(suffix, '')
        b = b.replace(suffix, '')
    a = a.strip().rstrip('.')
    b = b.strip().rstrip('.')

    if a == b:
        return True
    if a in b or b in a:
        return True
    # Check if first significant word matches
    a_words = a.split()
    b_words = b.split()
    if a_words and b_words and a_words[0] == b_words[0] and len(a_words[0]) > 3:
        return True
    return False


def _location_match(loc_a, loc_b):
    """Check if two normalized locations match."""
    if not loc_a or not loc_b:
        # If either is empty, consider it a broad match
        return not loc_a and not loc_b
    a = loc_a.lower().strip()
    b = loc_b.lower().strip()
    if a == b:
        return True
    if a in b or b in a:
        return True

    # Handle "san antonio/austin" matching either
    if "/" in a:
        parts = a.split("/")
        return any(p.strip() == b for p in parts)
    if "/" in b:
        parts = b.split("/")
        return any(p.strip() == a for p in parts)

    return False


# =============================================================================
# =============================================================================
# HTML GENERATION
# =============================================================================

LOCATION_TO_STATE = {
    'chicago': 'IL', 'bedford park': 'IL', 'hodgkins': 'IL', 'joliet': 'IL', 'libertyville': 'IL',
    'dallas': 'TX', 'fort worth': 'TX', 'lancaster': 'TX', 'haslet': 'TX', 'flower mound': 'TX',
    'plano': 'TX', 'arlington': 'TX', 'austin': 'TX', 'houston': 'TX', 'san antonio': 'TX',
    'nashville': 'TN', 'lavergne': 'TN', 'la vergne': 'TN', 'lebanon': 'TN',
    'las vegas': 'NV', 'reno': 'NV', 'mccarran': 'NV', 'sparks': 'NV',
    'atlanta': 'GA', 'hapeville': 'GA', 'cartersville': 'GA', 'mcdonough': 'GA',
    'orlando': 'FL', 'kissimmee': 'FL',
    'columbus': 'OH', 'grove city': 'OH', 'lockbourne': 'OH', 'cincinnati': 'OH',
    'middleburg heights': 'OH', 'west chester': 'OH', 'hebron': 'KY', 'erlanger': 'KY',
    'charlotte': 'NC', 'fort mill': 'SC',
    'phoenix': 'AZ',
    'washington, d.c.': 'DC', 'washington': 'DC',
    'logan township': 'NJ', 'paulsboro': 'NJ', 'south brunswick': 'NJ', 'newark': 'NJ',
    'philadelphia': 'PA', 'chester': 'PA',
    'inland empire': 'CA',
    'boston': 'MA',
    'fayetteville': 'AR',
}


def _get_state_from_location(location):
    """Get US state abbreviation from location string."""
    if not location:
        return 'Unknown'
    loc = location.lower().strip()
    # Direct match
    for key, state in LOCATION_TO_STATE.items():
        if key in loc:
            return state
    # Try to extract state from "City, ST" pattern
    m = re.search(r',\s*([A-Z]{2})\b', location)
    if m:
        return m.group(1)
    return 'Other'


def generate_html(processed_rows):
    """Generate the full HTML report."""
    # Sort by submission date (oldest first)
    processed_rows.sort(key=lambda r: r.get('submission_date') or date.max)
    all_states = set()
    all_roles = set()
    all_owners = set()

    # Calculate KPIs
    live_rows = [r for r in processed_rows if not r.get('is_declined') and not r.get('is_complete')]
    declined_rows = [r for r in processed_rows if r.get('is_declined')]
    complete_rows = [r for r in processed_rows if r.get('is_complete')]
    total_live = len(live_rows)
    total_declined = len(declined_rows)
    total_hc = sum(r['hc'] for r in live_rows if isinstance(r['hc'], int))
    no_fhs = sum(1 for r in live_rows if r['fhs']['count'] == 0)
    shifts_tbd = sum(1 for r in live_rows if not r['shifts'])

    rows_html = ""
    for r in processed_rows:
        is_declined = r.get('is_declined', False)
        is_complete = r.get('is_complete', False)

        if is_declined:
            bg = "#f3f4f6"  # Gray background for declined
        elif is_complete:
            bg = "#f9fafb"  # Light gray for complete
        else:
            bg = "#ffffff"

        # Status column: O (green) or C (black)
        if is_complete:
            status_cell = '<td style="text-align:center;font-weight:bold;font-size:14px;color:#fff;background:#111827;">C</td>'
        elif is_declined:
            status_cell = '<td style="text-align:center;font-weight:bold;font-size:14px;color:#fff;background:#6b7280;">D</td>'
        else:
            status_cell = '<td style="text-align:center;font-weight:bold;font-size:14px;color:#fff;background:#16a34a;">O</td>'

        hc_display = str(r['hc']) if isinstance(r['hc'], int) else r['hc']
        shifts_tag = ('<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:4px;font-size:12px;">Yes</span>'
                      if r['shifts'] else
                      '<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:4px;font-size:12px;">No</span>')

        if is_declined:
            status_label = '<span style="background:#9ca3af;color:white;padding:2px 6px;border-radius:4px;font-size:11px;">DECLINED</span>'
        elif is_complete:
            status_label = '<span style="background:#111827;color:white;padding:2px 6px;border-radius:4px;font-size:11px;">COMPLETE</span>'
        else:
            status_label = ""

        # Interview Target = HC × 10 (FHS section)
        hc_val = r['hc'] if isinstance(r['hc'], int) else 0
        int_target = hc_val * 10

        # RTB Target = HC × 2.5; Fill% = RTB ÷ (HC × 2.5)
        ob_rtb = r['ob']['rtb']
        rtb_target = round(hc_val * 2.5)
        fill_target = hc_val * 2.5
        if fill_target > 0:
            fill_pct = min((ob_rtb / fill_target) * 100, 100)
            fill_color = "#ef4444" if fill_pct < 25 else "#f59e0b" if fill_pct < 50 else "#22c55e"
            fill_html = f'''<div style="display:flex;align-items:center;gap:6px;">
        <div style="flex:1;background:#e5e7eb;border-radius:4px;height:14px;min-width:60px;">
            <div style="width:{fill_pct:.1f}%;background:{fill_color};height:100%;border-radius:4px;"></div>
        </div>
        <span style="font-size:13px;font-weight:bold;min-width:40px;text-align:right;">{fill_pct:.0f}%</span>
    </div>'''
        else:
            fill_html = '<span style="color:#9ca3af;">—</span>'

        # Escape HTML in role and client
        client_safe = r['client'].replace('&', '&amp;').replace('<', '&lt;')
        location_safe = r['location'].replace('&', '&amp;').replace('<', '&lt;')
        role_safe = r['role'].replace('&', '&amp;').replace('<', '&lt;') if r['role'] else ''

        # Build metric cells
        if is_declined or is_complete:
            client_suffix = f'<br>{status_label}'
        else:
            client_suffix = ''
        fhs_open_html = f'{r["fhs"]["count"]}'
        fhs_interview_html = f'{r["fhs"]["rsvps"]:,}'
        target_html = f'{int_target:,}' if int_target > 0 else '—'
        rtb_target_html = f'{int(rtb_target)}' if rtb_target > 0 else '—'
        ob_created_html = f'{r["ob"]["created"]}'
        ob_verified_html = f'{r["ob"]["verified"]}'
        ob_rtb_html = f'{r["ob"]["rtb"]}'
        fill_cell_html = fill_html

        # Format submission date for display
        sub_date = r.get('submission_date')
        sub_date_display = sub_date.strftime('%b/%d/%Y') if sub_date else '—'

        # Determine state from location
        row_state = _get_state_from_location(r['location'])
        all_states.add(row_state)

        # Collect role and owner for filters
        row_role = r.get('role', '').strip()
        row_owner = r.get('owner', '').strip()
        if row_role:
            all_roles.add(row_role)
        if row_owner:
            all_owners.add(row_owner)

        rows_html += f'''<tr style="background:{bg};" data-state="{row_state}" data-role="{row_role}" data-owner="{row_owner}">
        {status_cell}
        <td style="text-align:center;font-size:13px;">{sub_date_display}</td>
        <td style="text-align:center;font-size:13px;">{r['start_date']}</td>
        <td><strong>{client_safe}</strong><br><span style="color:#6b7280;font-size:12px;">{location_safe}</span>{client_suffix}</td>
        <td>{r['owner']}</td>
        <td>{role_safe}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{hc_display}</td>
        <td style="text-align:center;">{shifts_tag}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{fhs_open_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{fhs_interview_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{target_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_created_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_verified_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{ob_rtb_html}</td>
        <td style="text-align:center;font-weight:bold;font-size:15px;">{rtb_target_html}</td>
        <td style="min-width:120px;">{fill_cell_html}</td>
    </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Recruitment Request Dashboard — {REPORT_DATE}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f3f4f6; font-size: 14px; color: #111827; }}
        .header {{ background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%); color: white; padding: 24px 32px; }}
        .header h1 {{ font-size: 22px; font-weight: 700; }}
        .header p {{ font-size: 13px; opacity: 0.85; margin-top: 4px; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; padding: 20px 32px; }}
        .kpi-card {{ background: white; border-radius: 12px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); text-align: center; }}
        .kpi-card .number {{ font-size: 32px; font-weight: 800; line-height: 1.2; }}
        .kpi-card .label {{ font-size: 12px; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 4px; }}
        .kpi-card.red .number {{ color: #ef4444; }}
        .kpi-card.amber .number {{ color: #f59e0b; }}
        .kpi-card.blue .number {{ color: #2563eb; }}
        .kpi-card.green .number {{ color: #22c55e; }}
        .table-container {{ margin: 0 32px 32px; background: white; border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow: hidden; }}
        .table-scroll {{ overflow-y: auto; max-height: calc(100vh - 60px); }}
        thead tr:first-child th {{ position: sticky; top: 0; z-index: 20; }}
        thead tr:nth-child(2) th {{ position: sticky; top: 34px; z-index: 19; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
        th {{ position: sticky; top: 0; z-index: 10; background: #1e3a5f; color: white; padding: 8px 10px; font-size: 12px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.3px; white-space: nowrap; }}
        th[data-tip] {{ cursor: help; }}
        #tooltip {{
            position: fixed;
            background: #111827;
            color: #fff;
            padding: 10px 14px;
            border-radius: 8px;
            font-size: 13px;
            font-weight: 400;
            line-height: 1.5;
            max-width: 320px;
            z-index: 9999;
            box-shadow: 0 4px 16px rgba(0,0,0,0.4);
            pointer-events: none;
            display: none;
        }}
        th.group-header {{ background: #0f2540; padding: 6px 10px; font-size: 11px; text-align: center; border-left: 2px solid rgba(255,255,255,0.15); }}
        th.sub-header {{ background: #1e3a5f; font-size: 11px; }}
        th:first-child {{ border-left: none; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #e5e7eb; font-size: 14px; vertical-align: middle; }}
        tr:hover {{ filter: brightness(0.97); }}
        .legend {{ margin: 12px 32px; display: flex; gap: 24px; font-size: 12px; color: #6b7280; }}
        .legend span {{ display: inline-flex; align-items: center; gap: 6px; }}
        .legend .dot {{ width: 12px; height: 12px; border-radius: 3px; display: inline-block; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Recruitment Request Dashboard</h1>
        <p>Generated {REPORT_DATE} | Every Live revenue request = one row (no merging) | Data: FHS Requisitions, Indeed Campaign Report, OB Funnel</p>
    </div>

    <div class="kpi-grid">
        <div class="kpi-card blue">
            <div class="number">{total_live}</div>
            <div class="label">Live Requests</div>
        </div>
        <div class="kpi-card blue">
            <div class="number">{total_hc:,}</div>
            <div class="label">Total HC</div>
        </div>
        <div class="kpi-card red">
            <div class="number">{no_fhs}</div>
            <div class="label">No FHS</div>
        </div>
        <div class="kpi-card amber">
            <div class="number">{shifts_tbd}</div>
            <div class="label">Shifts TBD</div>
        </div>
    </div>

    <div class="legend">
        <span><span class="dot" style="background:#f3f4f6;border:1px solid #d1d5db;"></span> Declined / Complete</span>
        <span>Ordered by revenue team submission date</span>
    </div>

    <div style="margin: 12px 32px; display:flex; align-items:center; gap:12px; flex-wrap:wrap;">
        <label style="font-size:13px; font-weight:600; color:#374151;">State:</label>
        <select id="stateFilter" onchange="applyFilters()" style="padding:6px 12px; border-radius:6px; border:1px solid #d1d5db; font-size:13px; background:white; cursor:pointer;">
            <option value="all">All States</option>
            {''.join(f'<option value="{s}">{s}</option>' for s in sorted(all_states))}
        </select>
        <label style="font-size:13px; font-weight:600; color:#374151; margin-left:8px;">Role:</label>
        <select id="roleFilter" onchange="applyFilters()" style="padding:6px 12px; border-radius:6px; border:1px solid #d1d5db; font-size:13px; background:white; cursor:pointer;">
            <option value="all">All Roles</option>
            {''.join(f'<option value="{r}">{r}</option>' for r in sorted(all_roles))}
        </select>
        <label style="font-size:13px; font-weight:600; color:#374151; margin-left:8px;">Owner:</label>
        <select id="ownerFilter" onchange="applyFilters()" style="padding:6px 12px; border-radius:6px; border:1px solid #d1d5db; font-size:13px; background:white; cursor:pointer;">
            <option value="all">All Owners</option>
            {''.join(f'<option value="{o}">{o}</option>' for o in sorted(all_owners))}
        </select>
        <span id="rowCount" style="font-size:12px; color:#6b7280;"></span>
    </div>

    <div class="table-container">
        <div class="table-scroll">
            <table id="mainTable">
                <thead>
                    <tr>
                        <th rowspan="2" style="min-width:30px;" data-tip="O = Open (Live request), C = Closed (Completed)">St</th>
                        <th rowspan="2" style="min-width:90px;" data-tip="Date the revenue team submitted this request">Submitted</th>
                        <th rowspan="2" style="min-width:90px;" data-tip="Date when shifts start for this request">Shift Start</th>
                        <th rowspan="2" style="min-width:160px;" data-tip="Client name and location from the revenue request">Client</th>
                        <th rowspan="2" style="min-width:70px;" data-tip="Market owner responsible for this location">Owner</th>
                        <th rowspan="2" style="min-width:120px;" data-tip="Job role requested by the revenue team">Role</th>
                        <th rowspan="2" style="min-width:40px;" data-tip="Headcount — number of workers the revenue team needs to fill">HC</th>
                        <th rowspan="2" style="min-width:55px;" data-tip="Whether shift schedules have been posted for this request">Shifts</th>
                        <th colspan="3" class="group-header" data-tip="Data from FHS (Flex Hiring System) requisitions">FHS Requisitions</th>
                        <th colspan="3" class="group-header" data-tip="Data from the OB (Onboarding) Funnel — worker pipeline progress">OB Funnel</th>
                        <th colspan="2" class="group-header" data-tip="Fulfillment progress: RTB vs pipeline target">Fulfillment</th>
                    </tr>
                    <tr>
                        <th class="sub-header" style="text-align:center;" data-tip="Count of FHS requisitions with status Open or Auto-Paused for this client+location+role">Open</th>
                        <th class="sub-header" style="text-align:center;" data-tip="Sum of RSVPs (interview candidates) from FHS reqs created after the revenue request submission date">Interview</th>
                        <th class="sub-header" style="text-align:center;" data-tip="Target interviews needed = HC × 10 (10 interviews per hire)">Target</th>
                        <th class="sub-header" style="text-align:center;" data-tip="Worker Accounts Created — new workers who signed up via the OB funnel">Created</th>
                        <th class="sub-header" style="text-align:center;" data-tip="1st Role Verified — workers who completed role verification in the OB funnel">Verified</th>
                        <th class="sub-header" style="text-align:center;" data-tip="Ready to Book — workers who completed onboarding and are available for shifts">RTB</th>
                        <th class="sub-header" style="text-align:center;" data-tip="RTB Target = HC × 2.5 — pipeline target of Ready to Book workers needed">Target</th>
                        <th class="sub-header" style="text-align:center;" data-tip="Fill% = RTB ÷ (HC × 2.5) — workers ready to book vs. pipeline target">Fill%</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>
    </div>
    <div id="tooltip"></div>
    <script>
    // Tooltip logic
    const tip = document.getElementById('tooltip');
    document.querySelectorAll('th[data-tip]').forEach(th => {{
        th.addEventListener('mouseenter', e => {{
            tip.textContent = th.getAttribute('data-tip');
            tip.style.display = 'block';
            const rect = th.getBoundingClientRect();
            let left = rect.left + rect.width / 2 - tip.offsetWidth / 2;
            let top = rect.bottom + 8;
            // Keep within viewport
            if (left + tip.offsetWidth > window.innerWidth - 16) left = window.innerWidth - tip.offsetWidth - 16;
            if (left < 16) left = 16;
            if (top + tip.offsetHeight > window.innerHeight - 16) top = rect.top - tip.offsetHeight - 8;
            tip.style.left = left + 'px';
            tip.style.top = top + 'px';
        }});
        th.addEventListener('mouseleave', () => {{ tip.style.display = 'none'; }});
    }});

    function applyFilters() {{
        const stateVal = document.getElementById('stateFilter').value;
        const roleVal = document.getElementById('roleFilter').value;
        const ownerVal = document.getElementById('ownerFilter').value;
        const rows = document.querySelectorAll('#mainTable tbody tr');
        let visible = 0;
        const anyFilter = stateVal !== 'all' || roleVal !== 'all' || ownerVal !== 'all';
        rows.forEach(row => {{
            const matchState = stateVal === 'all' || row.dataset.state === stateVal;
            const matchRole = roleVal === 'all' || row.dataset.role === roleVal;
            const matchOwner = ownerVal === 'all' || row.dataset.owner === ownerVal;
            if (matchState && matchRole && matchOwner) {{
                row.style.display = '';
                visible++;
            }} else {{
                row.style.display = 'none';
            }}
        }});
        document.getElementById('rowCount').textContent = anyFilter ? visible + ' of ' + rows.length + ' rows' : '';
    }}
    </script>
</body>
</html>'''

    return html


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(processed_rows):
    """Generate an editable Excel workbook with the same data as the HTML dashboard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Sort by submission date
    processed_rows.sort(key=lambda r: r.get('submission_date') or date.max)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Requests"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    group_fill = PatternFill(start_color="0F2540", end_color="0F2540", fill_type="solid")
    declined_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    green_font = Font(color="166534", bold=True)
    red_font = Font(color="DC2626", bold=True)
    bold_font = Font(bold=True, size=11)
    number_font = Font(bold=True, size=12)
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Column definitions
    columns = [
        ('Submitted', 14),
        ('Shift Start', 14),
        ('Client', 22),
        ('Location', 18),
        ('Owner', 10),
        ('Role', 22),
        ('HC', 8),
        ('Shifts', 8),
        ('Status', 10),
        # FHS
        ('Open', 8),
        ('Interview', 11),
        ('Target', 10),
        # OB Funnel
        ('Created', 10),
        ('Verified', 10),
        ('RTB', 8),
        # Fill
        ('RTB Target', 10),
        ('Fill%', 10),
    ]

    # Single header row — no merged cells for Google Sheets compatibility
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows
    for row_idx, r in enumerate(processed_rows, 2):
        is_declined = r.get('is_declined', False)

        # Submission date
        sub_date = r.get('submission_date')
        ws.cell(row=row_idx, column=1, value=sub_date.strftime('%b/%d/%Y') if sub_date else '').alignment = center_align

        # Shift start
        ws.cell(row=row_idx, column=2, value=r['start_date']).alignment = center_align

        # Client
        ws.cell(row=row_idx, column=3, value=r['client']).font = bold_font

        # Location
        ws.cell(row=row_idx, column=4, value=r['location'])

        # Owner
        ws.cell(row=row_idx, column=5, value=r['owner'])

        # Role
        ws.cell(row=row_idx, column=6, value=r['role']).alignment = wrap_align

        # HC
        hc_val = r['hc']
        ws.cell(row=row_idx, column=7, value=hc_val if isinstance(hc_val, int) else str(hc_val)).alignment = center_align

        # Shifts
        ws.cell(row=row_idx, column=8, value='Yes' if r['shifts'] else 'No').alignment = center_align
        if r['shifts']:
            ws.cell(row=row_idx, column=8).font = green_font
        else:
            ws.cell(row=row_idx, column=8).font = red_font

        # Status
        ws.cell(row=row_idx, column=9, value='DECLINED' if is_declined else 'Live').alignment = center_align

        if is_declined:
            # Gray fill for entire row
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).fill = declined_fill

        # FHS Open
        ws.cell(row=row_idx, column=10, value=r['fhs']['count']).alignment = center_align
        ws.cell(row=row_idx, column=10).font = number_font

        # Interview
        ws.cell(row=row_idx, column=11, value=r['fhs']['rsvps']).alignment = center_align
        ws.cell(row=row_idx, column=11).font = number_font

        # Target (HC × 10)
        hc_num = r['hc'] if isinstance(r['hc'], int) else 0
        target = hc_num * 10
        ws.cell(row=row_idx, column=12, value=target if target > 0 else '').alignment = center_align
        ws.cell(row=row_idx, column=12).font = number_font

        # OB Created
        ws.cell(row=row_idx, column=13, value=r['ob']['created']).alignment = center_align
        ws.cell(row=row_idx, column=13).font = number_font

        # OB Verified
        ws.cell(row=row_idx, column=14, value=r['ob']['verified']).alignment = center_align
        ws.cell(row=row_idx, column=14).font = number_font

        # OB RTB
        ws.cell(row=row_idx, column=15, value=r['ob']['rtb']).alignment = center_align
        ws.cell(row=row_idx, column=15).font = number_font

        # RTB Target = HC × 2.5
        ws.cell(row=row_idx, column=16, value=round(hc_num * 2.5)).alignment = center_align
        ws.cell(row=row_idx, column=16).font = number_font

        # Fill% = RTB / (HC × 2.5)
        ob_rtb = r['ob']['rtb']
        fill_target = hc_num * 2.5
        if fill_target > 0:
            fill_pct = min(ob_rtb / fill_target, 1.0)
            cell = ws.cell(row=row_idx, column=17, value=fill_pct)
            cell.number_format = '0%'
            cell.alignment = center_align
            cell.font = number_font
            if fill_pct < 0.25:
                cell.font = Font(bold=True, size=12, color="DC2626")
            elif fill_pct < 0.50:
                cell.font = Font(bold=True, size=12, color="D97706")
            else:
                cell.font = Font(bold=True, size=12, color="16A34A")
        else:
            ws.cell(row=row_idx, column=17, value='').alignment = center_align

        # Bottom border for all cells
        for col in range(1, 18):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:Q{len(processed_rows) + 1}"

    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("Loading Revenue Requests...")
    revenue_rows = load_revenue_requests()
    live_count = sum(1 for r in revenue_rows if not r['is_declined'])
    declined_count = sum(1 for r in revenue_rows if r['is_declined'])
    print(f"  Found {live_count} Live + {declined_count} Declined = {len(revenue_rows)} total rows")

    print("Loading FHS Requisitions...")
    fhs_rows = load_fhs_requisitions()
    print(f"  Loaded {len(fhs_rows)} individual FHS rows")

    print("Loading Indeed Campaigns...")
    indeed_data = load_indeed_campaigns()
    print(f"  Loaded {len(indeed_data)} individual campaign rows")

    print("Loading OB Funnel...")
    ob_data = load_ob_funnel()
    print(f"  Loaded {len(ob_data)} client+location groups")

    # Load frozen snapshots and check for --refresh-frozen flag
    refresh_frozen = "--refresh-frozen" in sys.argv
    frozen = load_frozen_data()
    frozen_used = 0
    frozen_new = 0

    print("\nCross-referencing...")
    processed = []
    for i, row in enumerate(revenue_rows):
        key = _frozen_key(row)

        # Complete rows: use frozen snapshot if available (unless --refresh-frozen)
        if row['is_complete'] and key in frozen and not refresh_frozen:
            snap = frozen[key]
            fhs = snap['fhs']
            indeed = snap['indeed']
            ob = snap['ob']
            frozen_used += 1
            src = "FROZEN"
        else:
            # Live cross-reference
            fhs = match_fhs(row['client'], row['location'], fhs_rows, row.get('submission_date'), row.get('role'))
            indeed = match_indeed(row['client'], row['location'], indeed_data, row.get('role'))
            ob = match_ob_funnel(row['client'], row['location'], ob_data, row.get('role'))
            src = "LIVE"

            # Freeze newly-Complete rows
            if row['is_complete']:
                frozen[key] = {
                    'fhs': fhs,
                    'indeed': indeed,
                    'ob': ob,
                    'frozen_date': REPORT_DATE,
                }
                frozen_new += 1
                src = "FROZEN-NEW"

        processed.append({
            **row,
            'fhs': fhs,
            'indeed': indeed,
            'ob': ob,
        })

        print(f"  [{i+1:2d}] {row['client']:20s} | {row['location']:20s} | FHS={fhs['count']} | Indeed={indeed['count']} | OB Created={ob['created']} [{src}]")

    # Save updated frozen data
    save_frozen_data(frozen)
    print(f"  Frozen snapshots: {frozen_used} reused, {frozen_new} newly frozen, {len(frozen)} total stored")

    print(f"\nGenerating Excel report...")
    xlsx_path = generate_excel(processed)
    print(f"Excel saved to: {xlsx_path}")

    print(f"\nGenerating HTML report...")
    html = generate_html(processed)

    os.makedirs(os.path.dirname(OUTPUT_HTML), exist_ok=True)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"HTML saved to: {OUTPUT_HTML}")
    print(f"Total rows: {len(processed)}")

    # KPI summary
    total_hc = sum(r['hc'] for r in processed if isinstance(r['hc'], int))
    no_fhs = sum(1 for r in processed if r['fhs']['count'] == 0)
    shifts_tbd = sum(1 for r in processed if not r['shifts'])
    print(f"KPIs: Live={len(processed)}, HC={total_hc}, NoFHS={no_fhs}, ShiftsTBD={shifts_tbd}")


if __name__ == "__main__":
    main()
