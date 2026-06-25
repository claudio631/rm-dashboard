#!/usr/bin/env python3
"""
EOD Update Key Clients — Daily Slack Report Generator

Reads 3 data files from ~/Downloads and generates a Slack-formatted report.
Also copies the output to clipboard (macOS) for easy pasting.

Usage:
    python3 src/services/eod-report-generator.py
    python3 src/services/eod-report-generator.py --downloads-dir /path/to/downloads

Data files (auto-detected from ~/Downloads, most recent used):
    1. OB Funnel Custom Viewer.xlsx  — Tableau OB funnel split report
    2. JobsCampaigns_*.csv           — Indeed Analytics campaign spend
    3. requisitions-*.csv            — FHS Requisition report
"""

import csv
import glob
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

try:
    from google.ads.googleads.client import GoogleAdsClient
    GOOGLE_ADS_AVAILABLE = True
except ImportError:
    GOOGLE_ADS_AVAILABLE = False

GOOGLE_ADS_YAML = str(Path(__file__).resolve().parent.parent.parent / "google-ads.yaml")
US_ACCOUNT = "7236100723"


# --- Configuration ---

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
SPEND_HISTORY_FILE = PROJECT_ROOT / "src" / "data" / "last-report-spend.json"
DEFAULT_DOWNLOADS = Path.home() / "Downloads"

# Key clients and their named locations (order matters for output)
KEY_CLIENTS = {
    "CORT": {
        "named": ["Las Vegas", "Chicago", "Atlanta", "Orlando", "Phoenix", "Austin", "Nashville"],
        "group_others": True,
    },
    "Stord, Inc": {
        "named": ["Las Vegas", "Reno", "Atlanta", "Erlanger"],
        "group_others": False,
    },
    "OnTrac Final Mile": {
        "named": ["Logan Township", "Columbus", "Middleburg Heights", "Reno"],
        "group_others": True,
    },
    "CTDI": {
        "named": ["Dallas", "Columbus", "Nashville"],
        "group_others": False,
    },
}

CLIENT_ORDER = ["CORT", "Stord, Inc", "OnTrac Final Mile", "CTDI"]


# --- File Discovery ---

def find_latest_file(downloads_dir: Path, pattern: str) -> Optional[Path]:
    matches = sorted(
        downloads_dir.glob(pattern),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def discover_files(downloads_dir: Path) -> dict:
    ob_funnel = find_latest_file(downloads_dir, "OB Funnel Custom Viewer*.xlsx")
    jobs_campaigns = find_latest_file(downloads_dir, "JobsCampaigns_*.csv")
    requisitions = find_latest_file(downloads_dir, "requisitions-*.csv")

    missing = []
    if not ob_funnel:
        missing.append("OB Funnel Custom Viewer.xlsx")
    if not jobs_campaigns:
        missing.append("JobsCampaigns_*.csv")
    if not requisitions:
        missing.append("requisitions-*.csv")

    if missing:
        print(f"Error: Could not find the following files in {downloads_dir}:")
        for m in missing:
            print(f"  - {m}")
        sys.exit(1)

    return {
        "ob_funnel": ob_funnel,
        "jobs_campaigns": jobs_campaigns,
        "requisitions": requisitions,
    }


# --- Section 1: Key Client Unique Accounts ---

def get_comparison_date(today: datetime) -> datetime:
    """D-1 for Tue-Fri, D-3 (Friday) for Monday."""
    if today.weekday() == 0:  # Monday
        return today - timedelta(days=3)  # Friday
    return today - timedelta(days=1)


def process_ob_funnel(filepath: Path, today: datetime) -> dict:
    """Parse OB Funnel xlsx and return client/location data."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse header dates
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    date_cols = {}
    for i in range(3, len(header)):
        if header[i] and isinstance(header[i], str) and "'" in header[i]:
            try:
                d = datetime.strptime(header[i], "%b %d '%y")
                date_cols[i] = d
            except ValueError:
                pass

    window_start = today - timedelta(days=30)
    comparison_date = get_comparison_date(today)

    # Column indices
    window_cols = [i for i, d in date_cols.items() if window_start <= d <= today]
    comp_col = next((i for i, d in date_cols.items() if d == comparison_date), None)

    # Parse rows
    # Structure: col0=client, col1=location, col2=segment(Total/role), col3=metric, col4+=dates
    current_client = None
    current_location = None
    current_segment = None  # carries forward: "Total", "Loader / Crew", etc.
    results = {}
    # Track first-occurrence per (client, location, metric) to skip second Total blocks
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if vals[0]:
            current_client = vals[0]
        if vals[1]:
            current_location = vals[1]
        if vals[2] is not None:
            current_segment = vals[2]  # update when non-None: "Total", "Loader / Crew", etc.
        if not current_client or not current_location:
            continue

        if current_segment != "Total":
            continue

        metric = vals[3]  # actual metric name
        if metric not in ("Worker Accounts Created", "1st Role Verified (# Workers)"):
            continue

        dedup_key = (current_client, current_location, metric)
        if dedup_key in seen:
            continue  # skip second Total block for same (client, location, metric)
        seen.add(dedup_key)

        key = (current_client, current_location)
        if key not in results:
            results[key] = {"created": 0, "verified": 0, "delta": 0}

        if metric == "Worker Accounts Created":
            total = sum(
                int(vals[c]) for c in window_cols
                if c < len(vals) and vals[c] and isinstance(vals[c], (int, float))
            )
            results[key]["created"] = total

        elif metric == "1st Role Verified (# Workers)":
            total = sum(
                int(vals[c]) for c in window_cols
                if c < len(vals) and vals[c] and isinstance(vals[c], (int, float))
            )
            results[key]["verified"] = total

            d1_val = 0
            if comp_col and comp_col < len(vals) and vals[comp_col]:
                d1_val = int(vals[comp_col])
            results[key]["delta"] = d1_val

    wb.close()
    return results


def format_section1(data: dict) -> str:
    """Format key client unique accounts section."""
    lines = []

    for client_name in CLIENT_ORDER:
        config = KEY_CLIENTS[client_name]
        named_locs = config["named"]
        group_others = config["group_others"]

        # Collect all locations for this client
        client_data = [
            {"location": loc, **vals}
            for (c, loc), vals in data.items()
            if c == client_name
        ]

        if not client_data:
            continue

        lines.append(f"*{client_name}:*")

        # Named locations sorted by Created desc
        named_entries = []
        other_created = 0
        other_verified = 0
        other_delta = 0

        for loc_name in named_locs:
            entry = next((d for d in client_data if d["location"] == loc_name), None)
            if entry:
                named_entries.append(entry)

        # Sort named by created desc
        named_entries.sort(key=lambda x: -x["created"])

        for entry in named_entries:
            lines.append(
                f"{entry['location']}: {entry['created']} Created → "
                f"{entry['verified']} Verified (+{entry['delta']})"
            )

        # Other locations
        for entry in client_data:
            if entry["location"] not in named_locs:
                if group_others:
                    other_created += entry["created"]
                    other_verified += entry["verified"]
                    other_delta += entry["delta"]
                else:
                    # Show individually (sorted by created desc)
                    pass

        if not group_others:
            # Show non-named locations individually
            extras = sorted(
                [e for e in client_data if e["location"] not in named_locs],
                key=lambda x: -x["created"],
            )
            for entry in extras:
                lines.append(
                    f"{entry['location']}: {entry['created']} Created → "
                    f"{entry['verified']} Verified (+{entry['delta']})"
                )

        if group_others and (other_created > 0 or other_verified > 0):
            lines.append(
                f"Other Locations: {other_created} Created → "
                f"{other_verified} Verified (+{other_delta})"
            )

        lines.append("")  # blank line between clients

    return "\n".join(lines)


# --- Section 2: Indeed Spend ---

def process_indeed_spend(filepath: Path) -> float:
    """Sum the Spend column from JobsCampaigns CSV, current month only.

    Parses start/end dates from the filename:
      JobsCampaigns_YYYYMMDD_YYYYMMDD.csv

    - If both dates are in a prior month → file has no current-month data → return 0.
    - If the file starts in the current month → all rows are MTD → sum everything.
    - If the file spans months (start = prior, end = current) → filter by campaign
      name: include only campaigns whose first 8-digit date is in the current month,
      plus campaigns with no date in the name.
    """
    import re as _re
    today = datetime.now()
    current_ym = today.strftime("%Y%m")  # e.g. "202606"

    # Parse start and end dates from filename
    fname = filepath.stem  # e.g. "JobsCampaigns_20260501_20260526"
    m = _re.match(r"JobsCampaigns_(\d{6})\d{2}_(\d{6})\d{2}", fname)
    file_start_ym = m.group(1) if m else None   # "202605"
    file_end_ym   = m.group(2) if m else None   # "202605"

    # File is entirely in a prior month — no current-month data at all
    if file_end_ym and file_end_ym < current_ym:
        print(f"Warning: {filepath.name} only covers through {file_end_ym[:4]}-{file_end_ym[4:]} "
              f"— no {today.strftime('%B %Y')} data. Indeed June spend shown as $0.00. "
              f"Download a JobsCampaigns_{current_ym}01_*.csv for accurate figures.")
        return 0.0

    is_current_month_file = file_start_ym == current_ym

    # Pattern to find the first 8-digit date in a campaign name
    date_in_name = _re.compile(r"(\d{8})")

    total = 0.0
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not is_current_month_file:
                # Spans months: only include campaigns whose first 8-digit date
                # falls in the current month. Campaigns with NO date in the name
                # are EXCLUDED — their spend spans the full file period, not just
                # the current month, so including them inflates MTD figures.
                campaign_name = row.get("Campaign Name", "") or row.get("Campaign", "") or ""
                dates = date_in_name.findall(campaign_name)
                if not dates:
                    continue  # no date → can't attribute to current month → skip
                first_date_ym = dates[0][:6]  # first 6 chars = YYYYMM
                if first_date_ym != current_ym:
                    continue  # started in a prior month → skip
            try:
                total += float(row.get("Spend", 0))
            except (ValueError, TypeError):
                pass
    return total



def save_current_spend(indeed_spend: float, google_ads_spend: Optional[float] = None):
    """Save current spend for next report's comparison, preserving existing fields."""
    SPEND_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    existing = {}
    if SPEND_HISTORY_FILE.exists():
        with open(SPEND_HISTORY_FILE, "r") as f:
            try:
                existing = json.load(f)
            except json.JSONDecodeError:
                pass
    existing["last_spend"] = indeed_spend
    existing["last_spend_month"] = datetime.now().strftime("%Y-%m")
    existing["updated_at"] = datetime.now().isoformat()
    if google_ads_spend is not None:
        existing["last_google_ads_spend"] = google_ads_spend
    with open(SPEND_HISTORY_FILE, "w") as f:
        json.dump(existing, f, indent=2)


def get_google_ads_spend_mtd() -> Optional[float]:
    """Pull MTD spend from Google Ads API for US account."""
    if not GOOGLE_ADS_AVAILABLE:
        print("Warning: google-ads library not installed. Skipping Google Ads spend.")
        return None
    if not Path(GOOGLE_ADS_YAML).exists():
        print(f"Warning: {GOOGLE_ADS_YAML} not found. Skipping Google Ads spend.")
        return None
    try:
        client = GoogleAdsClient.load_from_storage(GOOGLE_ADS_YAML)
        ga_service = client.get_service("GoogleAdsService")
        today = datetime.now()
        month_start = today.replace(day=1).strftime("%Y-%m-%d")
        today_str = today.strftime("%Y-%m-%d")
        query = f"""
            SELECT metrics.cost_micros
            FROM campaign
            WHERE segments.date BETWEEN '{month_start}' AND '{today_str}'
              AND campaign.status != 'REMOVED'
        """
        response = ga_service.search(customer_id=US_ACCOUNT, query=query)
        total_micros = sum(row.metrics.cost_micros for row in response)
        return total_micros / 1_000_000
    except Exception as e:
        print(f"Warning: Google Ads API error: {e}")
        return None


def format_section2(current_spend: float, last_spend: Optional[float],
                    google_spend: Optional[float], last_google_spend: Optional[float]) -> str:
    """Format Indeed and Google Ads spend lines."""
    month_name = datetime.now().strftime("%B")
    lines = []
    indeed_line = f"*Indeed Spend Comparison:* Indeed {month_name} so far: ${current_spend:,.2f}"
    if last_spend is not None:
        delta = current_spend - last_spend
        indeed_line += f" (+${delta:,.2f} since last report)"
    lines.append(indeed_line)

    if google_spend is not None:
        google_line = f"*Google Ads Spend:* Google Ads {month_name} so far: ${google_spend:,.2f}"
        if last_google_spend is not None:
            delta = google_spend - last_google_spend
            google_line += f" (+${delta:,.2f} since last report)"
        lines.append(google_line)

    return "\n".join(lines)



# --- Section 3: Open Campaigns ---

def process_requisitions(filepath: Path) -> dict:
    """Extract open campaigns from requisitions CSV."""
    open_reqs = {}
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["status"].strip().lower() != "open":
                continue

            client = row["client"].strip()

            # Exclude Indeed Flex
            if "indeed flex" in client.lower():
                continue

            # Normalize location
            location = row["location"].strip()
            location = re.sub(r"\s*,\s*", ", ", location)
            location = re.sub(r"\s+", " ", location)

            if client not in open_reqs:
                open_reqs[client] = set()
            open_reqs[client].add(location)

    return open_reqs


def format_section3(data: dict) -> str:
    """Format open campaigns section."""
    lines = ["*Open Campaigns (Status: Open Only)*", ""]
    for client in sorted(data.keys()):
        locations = [l for l in sorted(data[client]) if l.strip()]
        if not locations:
            continue
        lines.append(f"{client}: {'; '.join(locations)}")
        lines.append("")
    return "\n".join(lines).rstrip()


# --- Main ---

def generate_report(downloads_dir: Path) -> str:
    """Generate the full EOD report."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    print(f"Report date: {today.strftime('%A, %B %d, %Y')}")
    print(f"Comparison day (D-1): {get_comparison_date(today).strftime('%A, %B %d')}")
    print()

    # Discover files
    files = discover_files(downloads_dir)
    print(f"OB Funnel:       {files['ob_funnel'].name}")
    print(f"JobsCampaigns:   {files['jobs_campaigns'].name}")
    print(f"Requisitions:    {files['requisitions'].name}")
    print()

    # Section 1: Key Client Unique Accounts
    ob_data = process_ob_funnel(files["ob_funnel"], today)
    section1 = format_section1(ob_data)

    # Section 2: Indeed Spend + Google Ads
    current_spend = process_indeed_spend(files["jobs_campaigns"])
    spend_history = {}
    if SPEND_HISTORY_FILE.exists():
        with open(SPEND_HISTORY_FILE, "r") as f:
            try:
                spend_history = json.load(f)
            except json.JSONDecodeError:
                pass
    # Only use last_spend for delta if it's from the same calendar month
    current_month = datetime.now().strftime("%Y-%m")
    stored_month = spend_history.get("last_spend_month", "")
    last_spend = spend_history.get("last_spend") if stored_month == current_month else None
    last_google_spend = spend_history.get("last_google_ads_spend")

    print("Pulling Google Ads MTD spend from API...")
    google_spend = get_google_ads_spend_mtd()
    if google_spend is not None:
        print(f"Google Ads MTD spend: ${google_spend:,.2f}")
    else:
        print("Google Ads spend unavailable — skipping from report.")

    section2 = format_section2(current_spend, last_spend, google_spend, last_google_spend)

    # Section 3: Open Campaigns
    req_data = process_requisitions(files["requisitions"])
    section3 = format_section3(req_data)

    # Assemble report
    report = (
        "@here EOD RM Update Key client unique accounts: (last update yesterday)\n"
        "\n"
        f"{section1}"
        "\n"
        f"{section2}\n"
        "\n"
        f"{section3}\n"
    )

    # Save spend for next run
    save_current_spend(current_spend, google_spend)

    return report


def copy_to_clipboard(text: str):
    """Copy text to macOS clipboard."""
    try:
        process = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE)
        process.communicate(text.encode("utf-8"))
        return True
    except FileNotFoundError:
        return False


def main():
    downloads_dir = DEFAULT_DOWNLOADS

    # Allow custom downloads dir via --downloads-dir flag
    if "--downloads-dir" in sys.argv:
        idx = sys.argv.index("--downloads-dir")
        if idx + 1 < len(sys.argv):
            downloads_dir = Path(sys.argv[idx + 1])

    report = generate_report(downloads_dir)

    print("=" * 60)
    print("SLACK REPORT (copy below)")
    print("=" * 60)
    print()
    print(report)

    # Copy to clipboard on macOS
    if copy_to_clipboard(report):
        print("=" * 60)
        print("Copied to clipboard! Paste directly into Slack.")
    else:
        print("=" * 60)
        print("Could not copy to clipboard. Copy the text above manually.")


if __name__ == "__main__":
    main()
